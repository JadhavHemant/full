import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Create workbook
wb = openpyxl.Workbook()

# Define styles
header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
header_font = Font(color="FFFFFF", bold=True, size=11)
get_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
post_fill = PatternFill(start_color="87CEEB", end_color="87CEEB", fill_type="solid")
put_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
patch_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
delete_fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")
not_started_fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# ============================================
# SHEET 1: Complete API List
# ============================================
ws1 = wb.active
ws1.title = "Complete API List"

# Title
ws1.merge_cells('A1:H1')
ws1['A1'] = "COMPLETE API INVENTORY - ERP/CRM System"
ws1['A1'].font = Font(size=16, bold=True)
ws1['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws1.row_dimensions[1].height = 30

# Subtitle
ws1.merge_cells('A2:H2')
ws1['A2'] = "Generated: 2026-07-22 | Total APIs: 200+"
ws1['A2'].alignment = Alignment(horizontal='center', vertical='center')
ws1['A2'].font = Font(size=10, italic=True)

# Headers
headers = ['Module', 'API Endpoint', 'HTTP Method', 'Controller Function', 'Authentication', 'Status', 'Priority', 'Notes']
for col, header in enumerate(headers, 1):
    cell = ws1.cell(row=4, column=col, value=header)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = thin_border

# Complete API Data
api_data = [
    # ==================== AUTHENTICATION APIs ====================
    ["Authentication", "/api/auth/register", "POST", "register", "Public", "Complete", "High", "User registration with OTP"],
    ["Authentication", "/api/auth/login", "POST", "login", "Public", "Complete", "High", "Login with rate limiting"],
    ["Authentication", "/api/auth/refresh", "POST", "refreshToken", "Public", "Complete", "High", "Refresh access token"],
    ["Authentication", "/api/auth/forgot-password", "POST", "forgotPassword", "Public", "Complete", "High", "Password reset request"],
    ["Authentication", "/api/auth/reset-password", "POST", "resetPassword", "Public", "Complete", "High", "Reset password with token"],
    ["Authentication", "/api/auth/verify-email", "GET", "verifyEmail", "Public", "Complete", "High", "Email verification"],
    ["Authentication", "/api/auth/resend-verification", "POST", "resendVerification", "Public", "Complete", "Medium", "Resend verification email"],
    ["Authentication", "/api/auth/logout", "POST", "logout", "Required", "Complete", "High", "Logout user"],
    ["Authentication", "/api/auth/logout-all", "POST", "logoutAll", "Required", "Complete", "High", "Logout from all devices"],
    ["Authentication", "/api/auth/change-password", "POST", "changePassword", "Required", "Complete", "High", "Change password"],
    ["Authentication", "/api/auth/me", "GET", "getMe", "Required", "Complete", "High", "Get current user profile"],
    ["Authentication", "/api/auth/unlock/:userId", "POST", "unlockAccount", "Required", "Complete", "Medium", "Unlock user account"],
    ["Authentication", "/api/auth/login-history", "GET", "getLoginHistory", "Required", "Complete", "Medium", "Get login history"],
    ["Authentication", "/api/auth/login-history/failed", "GET", "getFailedLogins", "Required", "Complete", "Medium", "Get failed login attempts"],
    ["Authentication", "/api/auth/login-history/suspicious", "GET", "getSuspiciousLogins", "Required", "Complete", "Medium", "Get suspicious logins"],
    ["Authentication", "/api/auth/login-history/stats", "GET", "getLoginStats", "Required", "Complete", "Medium", "Get login statistics"],
    ["Authentication", "/api/auth/active-sessions", "GET", "getActiveSessions", "Required", "Complete", "Medium", "Get active sessions"],
    ["Authentication", "/api/auth/sessions/:tokenId", "DELETE", "revokeSession", "Required", "Complete", "Medium", "Revoke session"],

    # ==================== TOKEN APIs ====================
    ["Token Management", "/api/token/refresh-token", "POST", "refreshAccessToken", "Public", "Complete", "High", "Refresh access token"],
    ["Token Management", "/api/token/logout", "POST", "logout", "Required", "Complete", "High", "Logout and revoke tokens"],

    # ==================== USER APIs ====================
    ["User Management", "/api/users/login", "POST", "loginUser", "Public", "Complete", "High", "User login"],
    ["User Management", "/api/users/register/send-otp", "POST", "sendRegistrationOtp", "Public", "Complete", "High", "Send OTP for registration"],
    ["User Management", "/api/users/register", "POST", "registerUser", "Public", "Complete", "High", "Register new user"],
    ["User Management", "/api/users/profile", "GET", "getProfile", "Required", "Complete", "High", "Get user profile"],
    ["User Management", "/api/users/update", "PUT", "updateUser", "Required", "Complete", "High", "Update user profile"],
    ["User Management", "/api/users/getall/profiles", "GET", "getAllUsers", "Required", "Complete", "High", "Get all users"],
    ["User Management", "/api/users/superadmin/company", "GET", "getCompanies", "Required", "Complete", "High", "Get all companies"],
    ["User Management", "/api/users/admin/company", "GET", "adminGetCompanies", "Required", "Complete", "Medium", "Get companies (admin)"],
    ["User Management", "/api/users/toggle-delete/:id", "PUT", "toggleSoftDelete", "Required", "Complete", "High", "Soft delete user"],
    ["User Management", "/api/users/toggle-activate/:id", "PUT", "toggleActivation", "Required", "Complete", "High", "Activate/deactivate user"],
    ["User Management", "/api/users/toggle-flag/:id", "PUT", "toggleFlag", "Required", "Complete", "Medium", "Toggle user flag"],
    ["User Management", "/api/users/forgot-password", "POST", "forgotPassword", "Public", "Complete", "High", "Forgot password"],
    ["User Management", "/api/users/reset-password", "POST", "resetPassword", "Public", "Complete", "High", "Reset password"],
    ["User Management", "/api/users/org/hierarchy", "GET", "getOrgHierarchy", "Required", "Complete", "Medium", "Get organization hierarchy"],
    ["User Management", "/api/users/my-team", "GET", "getMyTeamHierarchy", "Required", "Complete", "Medium", "Get my team hierarchy"],
    ["User Management", "/api/users/direct-reports/:userId", "GET", "getDirectReports", "Required", "Complete", "Medium", "Get direct reports"],
    ["User Management", "/api/users/company/:companyId/org-chart", "GET", "getCompanyOrgChart", "Required", "Complete", "Medium", "Get company org chart"],
    ["User Management", "/api/users/:userId/record-summary", "GET", "getUserRecordSummary", "Required", "Complete", "Medium", "Get user record summary"],
    ["User Management", "/api/users/:userId/records", "GET", "getUserModuleRecords", "Required", "Complete", "Medium", "Get user module records"],

    # ==================== ROLE APIs ====================
    ["Role Management", "/api/roles/", "GET", "getRoles", "Required", "Complete", "High", "Get all roles"],
    ["Role Management", "/api/roles/:roleId/permissions", "GET", "getRolePermissions", "Required", "Complete", "High", "Get role permissions"],
    ["Role Management", "/api/roles/:roleId/permissions", "POST", "saveRolePermissions", "Required", "Complete", "High", "Save role permissions"],
    ["Role Management", "/api/roles/create", "POST", "createRole", "Required", "Complete", "High", "Create new role"],
    ["Role Management", "/api/roles/:id", "PUT", "updateRole", "Required", "Complete", "High", "Update role"],
    ["Role Management", "/api/roles/:id", "DELETE", "deleteRole", "Required", "Complete", "High", "Delete role"],

    # ==================== RBAC APIs ====================
    ["RBAC - Modules", "/api/rbac/modules", "GET", "getModules", "Required", "Complete", "High", "Get all modules"],
    ["RBAC - Modules", "/api/rbac/modules/:moduleId", "GET", "getModuleById", "Required", "Complete", "High", "Get module by ID"],
    ["RBAC - Modules", "/api/rbac/modules", "POST", "createModule", "Required", "Complete", "High", "Create module"],
    ["RBAC - Modules", "/api/rbac/modules/:moduleId", "PUT", "updateModule", "Required", "Complete", "High", "Update module"],
    ["RBAC - Modules", "/api/rbac/modules/:moduleId", "DELETE", "deleteModule", "Required", "Complete", "High", "Delete module"],

    ["RBAC - Permissions", "/api/rbac/permissions", "GET", "getAllPermissions", "Required", "Complete", "High", "Get all permissions"],
    ["RBAC - Permissions", "/api/rbac/permissions", "POST", "createPermission", "Required", "Complete", "High", "Create permission"],
    ["RBAC - Permissions", "/api/rbac/permissions/:permissionId", "DELETE", "deletePermission", "Required", "Complete", "High", "Delete permission"],

    ["RBAC - Menus", "/api/rbac/menus", "GET", "getMenus", "Required", "Complete", "High", "Get all menus"],
    ["RBAC - Menus", "/api/rbac/menus/my-menus", "GET", "getUserMenus", "Required", "Complete", "High", "Get current user's menus"],
    ["RBAC - Menus", "/api/rbac/menus", "POST", "createMenu", "Required", "Complete", "High", "Create menu"],
    ["RBAC - Menus", "/api/rbac/menus/:menuId", "PUT", "updateMenu", "Required", "Complete", "High", "Update menu"],
    ["RBAC - Menus", "/api/rbac/menus/:menuId", "DELETE", "deleteMenu", "Required", "Complete", "High", "Delete menu"],

    ["RBAC - User Roles", "/api/rbac/user-roles", "GET", "getUserRoles", "Required", "Complete", "High", "Get user roles"],
    ["RBAC - User Roles", "/api/rbac/user-roles", "POST", "assignRole", "Required", "Complete", "High", "Assign role to user"],
    ["RBAC - User Roles", "/api/rbac/user-roles/:userRoleId", "DELETE", "revokeRole", "Required", "Complete", "High", "Revoke user role"],

    ["RBAC - Role Permissions", "/api/rbac/roles/:roleId/permissions", "GET", "getRolePermissions", "Required", "Complete", "High", "Get role permissions"],
    ["RBAC - Role Permissions", "/api/rbac/roles/:roleId/permissions", "POST", "assignPermissionsToRole", "Required", "Complete", "High", "Assign permissions to role"],
    ["RBAC - Role Permissions", "/api/rbac/roles/:roleId/permissions/:permissionId", "DELETE", "revokePermissionFromRole", "Required", "Complete", "High", "Revoke permission from role"],

    ["RBAC - Menu Permissions", "/api/rbac/roles/:roleId/menu-permissions", "GET", "getMenuPermissions", "Required", "Complete", "High", "Get menu permissions"],
    ["RBAC - Menu Permissions", "/api/rbac/roles/:roleId/menu-permissions", "POST", "setMenuPermissions", "Required", "Complete", "High", "Set menu permissions"],

    ["RBAC - User Permissions", "/api/rbac/users/:userId/permissions", "GET", "getUserPermissionsSummary", "Required", "Complete", "High", "Get user permissions summary"],

    # ==================== COMPANY APIs ====================
    ["Company Management", "/api/companies/create", "POST", "createCompany", "Required", "Complete", "High", "Create company"],
    ["Company Management", "/api/companies/list", "GET", "getCompanies", "Required", "Complete", "High", "Get all companies"],
    ["Company Management", "/api/companies/active", "GET", "getActiveCompanies", "Public", "Complete", "High", "Get active companies"],
    ["Company Management", "/api/companies/stats", "GET", "getCompanyStats", "Required", "Complete", "Medium", "Get company statistics"],
    ["Company Management", "/api/companies/deleted", "GET", "getDeletedCompanies", "Required", "Complete", "Medium", "Get deleted companies"],
    ["Company Management", "/api/companies/export", "GET", "exportCompanies", "Required", "Complete", "Medium", "Export companies to CSV"],
    ["Company Management", "/api/companies/bulk-delete", "DELETE", "bulkDeleteCompanies", "Required", "Complete", "Medium", "Bulk delete companies"],
    ["Company Management", "/api/companies/:id", "GET", "getCompanyById", "Required", "Complete", "High", "Get company by ID"],
    ["Company Management", "/api/companies/:id", "PUT", "updateCompany", "Required", "Complete", "High", "Update company"],
    ["Company Management", "/api/companies/delete/:id", "DELETE", "softDeleteCompany", "Required", "Complete", "High", "Soft delete company"],
    ["Company Management", "/api/companies/:id/restore", "PATCH", "restoreCompany", "Required", "Complete", "Medium", "Restore deleted company"],
    ["Company Management", "/api/companies/:id/toggle-active", "PATCH", "toggleActiveCompany", "Required", "Complete", "Medium", "Toggle company active status"],
    ["Company Management", "/api/companies/:id/toggle-flag", "PATCH", "toggleFlagCompany", "Required", "Complete", "Medium", "Toggle company flag"],

    # ==================== INVENTORY - PRODUCT APIs ====================
    ["Inventory - Products", "/api/products/", "GET", "getAllProducts", "Required", "Complete", "High", "Get all products"],
    ["Inventory - Products", "/api/products/", "POST", "createProduct", "Required", "Complete", "High", "Create product"],
    ["Inventory - Products", "/api/products/create", "POST", "createProduct", "Required", "Complete", "High", "Create product (alias)"],
    ["Inventory - Products", "/api/products/list", "GET", "getAllProducts", "Required", "Complete", "High", "Get products with filters"],
    ["Inventory - Products", "/api/products/active", "GET", "getAllProducts", "Required", "Complete", "High", "Get active products"],
    ["Inventory - Products", "/api/products/alerts/low-stock", "GET", "getLowStockProducts", "Required", "Complete", "High", "Get low stock products"],
    ["Inventory - Products", "/api/products/reports/stats", "GET", "getProductStats", "Required", "Complete", "Medium", "Get product statistics"],
    ["Inventory - Products", "/api/products/bulk-delete", "DELETE", "bulkDeleteProducts", "Required", "Complete", "Medium", "Bulk delete products"],
    ["Inventory - Products", "/api/products/:id", "GET", "getProductById", "Required", "Complete", "High", "Get product by ID"],
    ["Inventory - Products", "/api/products/:id", "PUT", "updateProduct", "Required", "Complete", "High", "Update product"],
    ["Inventory - Products", "/api/products/:id/toggle-active", "PATCH", "toggleActiveStatus", "Required", "Complete", "Medium", "Toggle product active status"],
    ["Inventory - Products", "/api/products/delete/:id", "DELETE", "softDeleteProduct", "Required", "Complete", "High", "Soft delete product"],

    # ==================== INVENTORY - PRODUCT CATEGORY APIs ====================
    ["Inventory - Categories", "/api/product-categories/", "GET", "getAllCategories", "Required", "Complete", "High", "Get all categories"],
    ["Inventory - Categories", "/api/product-categories/", "POST", "createCategory", "Required", "Complete", "High", "Create category"],
    ["Inventory - Categories", "/api/product-categories/:id", "GET", "getCategoryById", "Required", "Complete", "High", "Get category by ID"],
    ["Inventory - Categories", "/api/product-categories/:id", "PUT", "updateCategory", "Required", "Complete", "High", "Update category"],
    ["Inventory - Categories", "/api/product-categories/:id", "DELETE", "deleteCategory", "Required", "Complete", "High", "Delete category"],

    # ==================== INVENTORY - UNIT APIs ====================
    ["Inventory - Units", "/api/units/", "GET", "getAllUnits", "Required", "Complete", "High", "Get all units"],
    ["Inventory - Units", "/api/units/", "POST", "createUnit", "Required", "Complete", "High", "Create unit"],
    ["Inventory - Units", "/api/units/:id", "GET", "getUnitById", "Required", "Complete", "High", "Get unit by ID"],
    ["Inventory - Units", "/api/units/:id", "PUT", "updateUnit", "Required", "Complete", "High", "Update unit"],
    ["Inventory - Units", "/api/units/:id", "DELETE", "deleteUnit", "Required", "Complete", "High", "Delete unit"],

    # ==================== INVENTORY - WAREHOUSE APIs ====================
    ["Inventory - Warehouses", "/api/warehouses/", "GET", "getAllWarehouses", "Required", "Complete", "High", "Get all warehouses"],
    ["Inventory - Warehouses", "/api/warehouses/", "POST", "createWarehouse", "Required", "Complete", "High", "Create warehouse"],
    ["Inventory - Warehouses", "/api/warehouses/:id", "GET", "getWarehouseById", "Required", "Complete", "High", "Get warehouse by ID"],
    ["Inventory - Warehouses", "/api/warehouses/:id", "PUT", "updateWarehouse", "Required", "Complete", "High", "Update warehouse"],
    ["Inventory - Warehouses", "/api/warehouses/:id", "DELETE", "deleteWarehouse", "Required", "Complete", "High", "Delete warehouse"],

    # ==================== INVENTORY - STOCK APIs ====================
    ["Inventory - Stock", "/api/stock/", "GET", "getAllStock", "Required", "Complete", "High", "Get all stock"],
    ["Inventory - Stock", "/api/stock/warehouse/:warehouseId", "GET", "getStockByWarehouse", "Required", "Complete", "High", "Get stock by warehouse"],
    ["Inventory - Stock", "/api/stock/product/:productId", "GET", "getStockByProduct", "Required", "Complete", "High", "Get stock by product"],

    # ==================== INVENTORY - STOCK MOVEMENT APIs ====================
    ["Inventory - Stock Movements", "/api/stock-movements/", "GET", "getAllStockMovements", "Required", "Complete", "High", "Get all stock movements"],
    ["Inventory - Stock Movements", "/api/stock-movements/", "POST", "createStockMovement", "Required", "Complete", "High", "Create stock movement"],
    ["Inventory - Stock Movements", "/api/stock-movements/:id", "GET", "getStockMovementById", "Required", "Complete", "High", "Get stock movement by ID"],

    # ==================== INVENTORY - SUPPLIER APIs ====================
    ["Inventory - Suppliers", "/api/suppliers/", "GET", "getAllSuppliers", "Required", "Complete", "High", "Get all suppliers"],
    ["Inventory - Suppliers", "/api/suppliers/", "POST", "createSupplier", "Required", "Complete", "High", "Create supplier"],
    ["Inventory - Suppliers", "/api/suppliers/:id", "GET", "getSupplierById", "Required", "Complete", "High", "Get supplier by ID"],
    ["Inventory - Suppliers", "/api/suppliers/:id", "PUT", "updateSupplier", "Required", "Complete", "High", "Update supplier"],
    ["Inventory - Suppliers", "/api/suppliers/:id", "DELETE", "deleteSupplier", "Required", "Complete", "High", "Delete supplier"],

    # ==================== INVENTORY - CUSTOMER APIs ====================
    ["Inventory - Customers", "/api/customers/", "GET", "getAllCustomers", "Required", "Complete", "High", "Get all customers"],
    ["Inventory - Customers", "/api/customers/", "POST", "createCustomer", "Required", "Complete", "High", "Create customer"],
    ["Inventory - Customers", "/api/customers/:id", "GET", "getCustomerById", "Required", "Complete", "High", "Get customer by ID"],
    ["Inventory - Customers", "/api/customers/:id", "PUT", "updateCustomer", "Required", "Complete", "High", "Update customer"],
    ["Inventory - Customers", "/api/customers/:id", "DELETE", "deleteCustomer", "Required", "Complete", "High", "Delete customer"],

    # ==================== PURCHASE ORDER APIs ====================
    ["Purchase Orders", "/api/purchase-orders/stats", "GET", "getPurchaseOrderStats", "Required", "Complete", "High", "Get PO statistics"],
    ["Purchase Orders", "/api/purchase-orders/supplier/:supplierId", "GET", "getPurchaseOrdersBySupplier", "Required", "Complete", "Medium", "Get POs by supplier"],
    ["Purchase Orders", "/api/purchase-orders/", "GET", "getAllPurchaseOrders", "Required", "Complete", "High", "Get all POs"],
    ["Purchase Orders", "/api/purchase-orders/:id", "GET", "getPurchaseOrderById", "Required", "Complete", "High", "Get PO by ID"],
    ["Purchase Orders", "/api/purchase-orders/", "POST", "createPurchaseOrder", "Required", "Complete", "High", "Create purchase order"],
    ["Purchase Orders", "/api/purchase-orders/:id", "PUT", "updatePurchaseOrder", "Required", "Complete", "High", "Update purchase order"],
    ["Purchase Orders", "/api/purchase-orders/:id/status", "PATCH", "updatePurchaseOrderStatus", "Required", "Complete", "High", "Update PO status"],
    ["Purchase Orders", "/api/purchase-orders/soft-delete/:id", "PATCH", "softDeletePurchaseOrder", "Required", "Complete", "Medium", "Soft delete PO"],
    ["Purchase Orders", "/api/purchase-orders/hard-delete/:id", "DELETE", "hardDeletePurchaseOrder", "Required", "Complete", "Medium", "Hard delete PO"],

    # ==================== PURCHASE ORDER ITEM APIs ====================
    ["Purchase Order Items", "/api/purchase-order-items/", "GET", "getAllPOItems", "Required", "Complete", "High", "Get all PO items"],
    ["Purchase Order Items", "/api/purchase-order-items/", "POST", "createPOItem", "Required", "Complete", "High", "Create PO item"],
    ["Purchase Order Items", "/api/purchase-order-items/:id", "GET", "getPOItemById", "Required", "Complete", "High", "Get PO item by ID"],
    ["Purchase Order Items", "/api/purchase-order-items/:id", "PUT", "updatePOItem", "Required", "Complete", "High", "Update PO item"],
    ["Purchase Order Items", "/api/purchase-order-items/:id", "DELETE", "deletePOItem", "Required", "Complete", "High", "Delete PO item"],

    # ==================== SALES ORDER APIs ====================
    ["Sales Orders", "/api/sales-orders/stats", "GET", "getSalesOrderStats", "Required", "Complete", "High", "Get SO statistics"],
    ["Sales Orders", "/api/sales-orders/customer/:customerId", "GET", "getSalesOrdersByCustomer", "Required", "Complete", "Medium", "Get SOs by customer"],
    ["Sales Orders", "/api/sales-orders/", "GET", "getAllSalesOrders", "Required", "Complete", "High", "Get all sales orders"],
    ["Sales Orders", "/api/sales-orders/:id", "GET", "getSalesOrderById", "Required", "Complete", "High", "Get SO by ID"],
    ["Sales Orders", "/api/sales-orders/", "POST", "createSalesOrder", "Required", "Complete", "High", "Create sales order"],
    ["Sales Orders", "/api/sales-orders/:id", "PUT", "updateSalesOrder", "Required", "Complete", "High", "Update sales order"],
    ["Sales Orders", "/api/sales-orders/:id/status", "PATCH", "updateSalesOrderStatus", "Required", "Complete", "High", "Update SO status"],
    ["Sales Orders", "/api/sales-orders/:id/payment", "PATCH", "updatePayment", "Required", "Complete", "High", "Update payment"],
    ["Sales Orders", "/api/sales-orders/:id/restore", "PATCH", "restoreSalesOrder", "Required", "Complete", "Medium", "Restore SO"],
    ["Sales Orders", "/api/sales-orders/:id/soft-delete", "PATCH", "softDeleteSalesOrder", "Required", "Complete", "Medium", "Soft delete SO"],
    ["Sales Orders", "/api/sales-orders/:id", "DELETE", "hardDeleteSalesOrder", "Required", "Complete", "Medium", "Hard delete SO"],

    # ==================== SALES ORDER ITEM APIs ====================
    ["Sales Order Items", "/api/sales-order-items/", "GET", "getAllSOItems", "Required", "Complete", "High", "Get all SO items"],
    ["Sales Order Items", "/api/sales-order-items/", "POST", "createSOItem", "Required", "Complete", "High", "Create SO item"],
    ["Sales Order Items", "/api/sales-order-items/:id", "GET", "getSOItemById", "Required", "Complete", "High", "Get SO item by ID"],
    ["Sales Order Items", "/api/sales-order-items/:id", "PUT", "updateSOItem", "Required", "Complete", "High", "Update SO item"],
    ["Sales Order Items", "/api/sales-order-items/:id", "DELETE", "deleteSOItem", "Required", "Complete", "High", "Delete SO item"],

    # ==================== GRN APIs ====================
    ["GRN", "/api/grn/grn", "POST", "createGRN", "Required", "Complete", "High", "Create GRN"],
    ["GRN", "/api/grn/grn", "GET", "getAllGRNs", "Required", "Complete", "High", "Get all GRNs"],
    ["GRN", "/api/grn/grn/:id", "GET", "getGRNById", "Required", "Complete", "High", "Get GRN by ID"],

    # ==================== STOCK TRANSFER APIs ====================
    ["Stock Transfers", "/api/stock-transfers/", "POST", "createStockTransfer", "Required", "Complete", "High", "Create stock transfer"],
    ["Stock Transfers", "/api/stock-transfers/", "GET", "getAllStockTransfers", "Required", "Complete", "High", "Get all stock transfers"],
    ["Stock Transfers", "/api/stock-transfers/:id", "GET", "getStockTransferById", "Required", "Complete", "High", "Get transfer by ID"],
    ["Stock Transfers", "/api/stock-transfers/:id/status", "PUT", "updateTransferStatus", "Required", "Complete", "High", "Update transfer status"],

    # ==================== STOCK ADJUSTMENT APIs ====================
    ["Stock Adjustments", "/api/stock-adjustments/", "GET", "getAllStockAdjustments", "Required", "Complete", "High", "Get all adjustments"],
    ["Stock Adjustments", "/api/stock-adjustments/", "POST", "createStockAdjustment", "Required", "Complete", "High", "Create adjustment"],
    ["Stock Adjustments", "/api/stock-adjustments/:id", "GET", "getStockAdjustmentById", "Required", "Complete", "High", "Get adjustment by ID"],

    # ==================== BATCH/SERIAL APIs ====================
    ["Batch/Serial Tracking", "/api/batch-serial/", "GET", "getAllBatchSerial", "Required", "Complete", "High", "Get all batch/serial"],
    ["Batch/Serial Tracking", "/api/batch-serial/", "POST", "createBatchSerial", "Required", "Complete", "High", "Create batch/serial"],
    ["Batch/Serial Tracking", "/api/batch-serial/:id", "GET", "getBatchSerialById", "Required", "Complete", "High", "Get by ID"],
    ["Batch/Serial Tracking", "/api/batch-serial/:id", "PUT", "updateBatchSerial", "Required", "Complete", "High", "Update batch/serial"],
    ["Batch/Serial Tracking", "/api/batch-serial/:id", "DELETE", "deleteBatchSerial", "Required", "Complete", "High", "Delete batch/serial"],

    # ==================== BRAND APIs ====================
    ["Brands", "/api/brands/", "GET", "getAllBrands", "Required", "Complete", "Medium", "Get all brands"],
    ["Brands", "/api/brands/", "POST", "createBrand", "Required", "Complete", "Medium", "Create brand"],
    ["Brands", "/api/brands/:id", "GET", "getBrandById", "Required", "Complete", "Medium", "Get brand by ID"],
    ["Brands", "/api/brands/:id", "PUT", "updateBrand", "Required", "Complete", "Medium", "Update brand"],
    ["Brands", "/api/brands/:id", "DELETE", "deleteBrand", "Required", "Complete", "Medium", "Delete brand"],

    # ==================== TAX APIs ====================
    ["Taxes", "/api/taxes/", "GET", "getAllTaxes", "Required", "Complete", "High", "Get all taxes"],
    ["Taxes", "/api/taxes/", "POST", "createTax", "Required", "Complete", "High", "Create tax"],
    ["Taxes", "/api/taxes/:id", "GET", "getTaxById", "Required", "Complete", "High", "Get tax by ID"],
    ["Taxes", "/api/taxes/:id", "PUT", "updateTax", "Required", "Complete", "High", "Update tax"],
    ["Taxes", "/api/taxes/:id", "DELETE", "deleteTax", "Required", "Complete", "High", "Delete tax"],

    # ==================== PRODUCT TAX MAP APIs ====================
    ["Product Tax Map", "/api/product-tax-map/", "GET", "getAllProductTaxMap", "Required", "Complete", "High", "Get all product tax mappings"],
    ["Product Tax Map", "/api/product-tax-map/", "POST", "createProductTaxMap", "Required", "Complete", "High", "Create product tax mapping"],
    ["Product Tax Map", "/api/product-tax-map/:id", "GET", "getProductTaxMapById", "Required", "Complete", "High", "Get mapping by ID"],
    ["Product Tax Map", "/api/product-tax-map/:id", "PUT", "updateProductTaxMap", "Required", "Complete", "High", "Update mapping"],
    ["Product Tax Map", "/api/product-tax-map/:id", "DELETE", "deleteProductTaxMap", "Required", "Complete", "High", "Delete mapping"],

    # ==================== AUDIT LOG APIs ====================
    ["Audit Logs", "/api/audit-logs/", "GET", "getAllAuditLogs", "Required", "Complete", "High", "Get all audit logs"],
    ["Audit Logs", "/api/audit-logs/:id", "GET", "getAuditLogById", "Required", "Complete", "Medium", "Get audit log by ID"],

    # ==================== PROFIT/LOSS REPORT APIs ====================
    ["Profit/Loss Reports", "/api/profit-loss-reports/", "GET", "getAllProfitLossReports", "Required", "Complete", "High", "Get all reports"],
    ["Profit/Loss Reports", "/api/profit-loss-reports/", "POST", "createProfitLossReport", "Required", "Complete", "High", "Create report"],
    ["Profit/Loss Reports", "/api/profit-loss-reports/:id", "GET", "getProfitLossReportById", "Required", "Complete", "High", "Get report by ID"],

    # ==================== EXPORT APIs ====================
    ["Export", "/api/export/", "GET", "exportData", "Required", "Complete", "Medium", "Export data to CSV/Excel"],

    # ==================== CRM - LEAD APIs ====================
    ["CRM - Leads", "/api/leads/", "GET", "getAllLeads", "Required", "Complete", "High", "Get all leads"],
    ["CRM - Leads", "/api/leads/", "POST", "createLead", "Required", "Complete", "High", "Create lead"],
    ["CRM - Leads", "/api/leads/:id", "GET", "getLeadById", "Required", "Complete", "High", "Get lead by ID"],
    ["CRM - Leads", "/api/leads/:id", "PUT", "updateLead", "Required", "Complete", "High", "Update lead"],
    ["CRM - Leads", "/api/leads/:id", "DELETE", "deleteLead", "Required", "Complete", "High", "Delete lead"],

    # ==================== CRM - OPPORTUNITY APIs ====================
    ["CRM - Opportunities", "/api/opportunities/", "GET", "getAllOpportunities", "Required", "Complete", "High", "Get all opportunities"],
    ["CRM - Opportunities", "/api/opportunities/", "POST", "createOpportunity", "Required", "Complete", "High", "Create opportunity"],
    ["CRM - Opportunities", "/api/opportunities/:id", "GET", "getOpportunityById", "Required", "Complete", "High", "Get opportunity by ID"],
    ["CRM - Opportunities", "/api/opportunities/:id", "PUT", "updateOpportunity", "Required", "Complete", "High", "Update opportunity"],
    ["CRM - Opportunities", "/api/opportunities/:id", "DELETE", "deleteOpportunity", "Required", "Complete", "High", "Delete opportunity"],

    # ==================== CRM - ACCOUNT APIs ====================
    ["CRM - Accounts", "/api/accounts/", "GET", "getAllAccounts", "Required", "Complete", "High", "Get all accounts"],
    ["CRM - Accounts", "/api/accounts/", "POST", "createAccount", "Required", "Complete", "High", "Create account"],
    ["CRM - Accounts", "/api/accounts/:id", "GET", "getAccountById", "Required", "Complete", "High", "Get account by ID"],
    ["CRM - Accounts", "/api/accounts/:id", "PUT", "updateAccount", "Required", "Complete", "High", "Update account"],
    ["CRM - Accounts", "/api/accounts/:id", "DELETE", "deleteAccount", "Required", "Complete", "High", "Delete account"],

    # ==================== CRM - CONTACT APIs ====================
    ["CRM - Contacts", "/api/contacts/", "GET", "getAllContacts", "Required", "Complete", "High", "Get all contacts"],
    ["CRM - Contacts", "/api/contacts/", "POST", "createContact", "Required", "Complete", "High", "Create contact"],
    ["CRM - Contacts", "/api/contacts/:id", "GET", "getContactById", "Required", "Complete", "High", "Get contact by ID"],
    ["CRM - Contacts", "/api/contacts/:id", "PUT", "updateContact", "Required", "Complete", "High", "Update contact"],
    ["CRM - Contacts", "/api/contacts/:id", "DELETE", "deleteContact", "Required", "Complete", "High", "Delete contact"],

    # ==================== CRM - ACTIVITY APIs ====================
    ["CRM - Activities", "/api/activities/", "GET", "getAllActivities", "Required", "Complete", "High", "Get all activities"],
    ["CRM - Activities", "/api/activities/", "POST", "createActivity", "Required", "Complete", "High", "Create activity"],
    ["CRM - Activities", "/api/activities/:id", "GET", "getActivityById", "Required", "Complete", "High", "Get activity by ID"],
    ["CRM - Activities", "/api/activities/:id", "PUT", "updateActivity", "Required", "Complete", "High", "Update activity"],
    ["CRM - Activities", "/api/activities/:id", "DELETE", "deleteActivity", "Required", "Complete", "High", "Delete activity"],

    # ==================== CRM - QUOTE APIs ====================
    ["CRM - Quotes", "/api/quotes/", "GET", "getAllQuotes", "Required", "Complete", "High", "Get all quotes"],
    ["CRM - Quotes", "/api/quotes/", "POST", "createQuote", "Required", "Complete", "High", "Create quote"],
    ["CRM - Quotes", "/api/quotes/:id", "GET", "getQuoteById", "Required", "Complete", "High", "Get quote by ID"],
    ["CRM - Quotes", "/api/quotes/:id", "PUT", "updateQuote", "Required", "Complete", "High", "Update quote"],
    ["CRM - Quotes", "/api/quotes/:id", "DELETE", "deleteQuote", "Required", "Complete", "High", "Delete quote"],

    # ==================== CRM - INVOICE APIs ====================
    ["CRM - Invoices", "/api/invoices/", "GET", "getAllInvoices", "Required", "Complete", "High", "Get all invoices"],
    ["CRM - Invoices", "/api/invoices/", "POST", "createInvoice", "Required", "Complete", "High", "Create invoice"],
    ["CRM - Invoices", "/api/invoices/:id", "GET", "getInvoiceById", "Required", "Complete", "High", "Get invoice by ID"],
    ["CRM - Invoices", "/api/invoices/:id", "PUT", "updateInvoice", "Required", "Complete", "High", "Update invoice"],
    ["CRM - Invoices", "/api/invoices/:id", "DELETE", "deleteInvoice", "Required", "Complete", "High", "Delete invoice"],

    # ==================== CRM - PAYMENT APIs ====================
    ["CRM - Payments", "/api/payments/", "GET", "getAllPayments", "Required", "Complete", "High", "Get all payments"],
    ["CRM - Payments", "/api/payments/", "POST", "createPayment", "Required", "Complete", "High", "Create payment"],
    ["CRM - Payments", "/api/payments/:id", "GET", "getPaymentById", "Required", "Complete", "High", "Get payment by ID"],
    ["CRM - Payments", "/api/payments/:id", "PUT", "updatePayment", "Required", "Complete", "High", "Update payment"],
    ["CRM - Payments", "/api/payments/:id", "DELETE", "deletePayment", "Required", "Complete", "High", "Delete payment"],

    # ==================== CRM - CASE APIs ====================
    ["CRM - Cases", "/api/cases/", "GET", "getAllCases", "Required", "Complete", "High", "Get all cases"],
    ["CRM - Cases", "/api/cases/", "POST", "createCase", "Required", "Complete", "High", "Create case"],
    ["CRM - Cases", "/api/cases/:id", "GET", "getCaseById", "Required", "Complete", "High", "Get case by ID"],
    ["CRM - Cases", "/api/cases/:id", "PUT", "updateCase", "Required", "Complete", "High", "Update case"],
    ["CRM - Cases", "/api/cases/:id", "DELETE", "deleteCase", "Required", "Complete", "High", "Delete case"],

    # ==================== CRM - RETENTION APIs ====================
    ["CRM - Retentions", "/api/retentions/", "GET", "getAllRetentions", "Required", "Complete", "Medium", "Get all retentions"],
    ["CRM - Retentions", "/api/retentions/", "POST", "createRetention", "Required", "Complete", "Medium", "Create retention"],
    ["CRM - Retentions", "/api/retentions/:id", "GET", "getRetentionById", "Required", "Complete", "Medium", "Get retention by ID"],
    ["CRM - Retentions", "/api/retentions/:id", "PUT", "updateRetention", "Required", "Complete", "Medium", "Update retention"],
    ["CRM - Retentions", "/api/retentions/:id", "DELETE", "deleteRetention", "Required", "Complete", "Medium", "Delete retention"],

    # ==================== CRM - PRESALES APIs ====================
    ["CRM - Presales", "/api/presales/", "GET", "getAllPresales", "Required", "Complete", "Medium", "Get all presales"],
    ["CRM - Presales", "/api/presales/", "POST", "createPresales", "Required", "Complete", "Medium", "Create presales"],
    ["CRM - Presales", "/api/presales/:id", "GET", "getPresalesById", "Required", "Complete", "Medium", "Get presales by ID"],
    ["CRM - Presales", "/api/presales/:id", "PUT", "updatePresales", "Required", "Complete", "Medium", "Update presales"],
    ["CRM - Presales", "/api/presales/:id", "DELETE", "deletePresales", "Required", "Complete", "Medium", "Delete presales"],

    # ==================== CRM - MASTER DATA APIs ====================
    ["CRM - Master Data", "/api/lead-sources/", "GET", "getAllLeadSources", "Required", "Complete", "High", "Get lead sources"],
    ["CRM - Master Data", "/api/lead-sources/", "POST", "createLeadSource", "Required", "Complete", "High", "Create lead source"],
    ["CRM - Master Data", "/api/industries/", "GET", "getAllIndustries", "Required", "Complete", "Medium", "Get industries"],
    ["CRM - Master Data", "/api/industries/", "POST", "createIndustry", "Required", "Complete", "Medium", "Create industry"],
    ["CRM - Master Data", "/api/sales-stages/", "GET", "getAllSalesStages", "Required", "Complete", "High", "Get sales stages"],
    ["CRM - Master Data", "/api/sales-stages/", "POST", "createSalesStage", "Required", "Complete", "High", "Create sales stage"],
    ["CRM - Master Data", "/api/followup-types/", "GET", "getAllFollowupTypes", "Required", "Complete", "Medium", "Get follow-up types"],
    ["CRM - Master Data", "/api/followup-types/", "POST", "createFollowupType", "Required", "Complete", "Medium", "Create follow-up type"],
    ["CRM - Master Data", "/api/task-types/", "GET", "getAllTaskTypes", "Required", "Complete", "Medium", "Get task types"],
    ["CRM - Master Data", "/api/task-types/", "POST", "createTaskType", "Required", "Complete", "Medium", "Create task type"],

    # ==================== SYSTEM - REPORTS APIs ====================
    ["System - Reports", "/api/reports/dashboard", "GET", "getSuperAdminDashboard", "Required", "Complete", "High", "Get admin dashboard"],
    ["System - Reports", "/api/reports/employee-activity", "GET", "getEmployeeActivity", "Required", "Complete", "Medium", "Get employee activity"],
    ["System - Reports", "/api/reports/notifications", "GET", "getRecentNotifications", "Required", "Complete", "Medium", "Get recent notifications"],
    ["System - Reports", "/api/reports/overview", "GET", "getReportOverview", "Required", "Complete", "High", "Get report overview"],
    ["System - Reports", "/api/reports/crm-digest/run", "POST", "triggerCrmDigestReport", "Required", "Complete", "Medium", "Trigger CRM digest report"],

    # ==================== SYSTEM - AUDIT APIs ====================
    ["System - Audit", "/api/audit-events/", "GET", "listAuditEvents", "Required", "Complete", "High", "List audit events"],
    ["System - Audit", "/api/audit-events/", "POST", "createAuditEvent", "Required", "Complete", "High", "Create audit event"],

    # ==================== SYSTEM - COMPANY SETTINGS APIs ====================
    ["System - Company Settings", "/api/company-settings/", "GET", "getCompanySettings", "Required", "Complete", "Medium", "Get company settings"],
    ["System - Company Settings", "/api/company-settings/:companyId", "GET", "getCompanySettings", "Required", "Complete", "Medium", "Get settings by company"],
    ["System - Company Settings", "/api/company-settings/", "PUT", "upsertCompanySettings", "Required", "Complete", "Medium", "Update company settings"],
    ["System - Company Settings", "/api/company-settings/:companyId", "PUT", "upsertCompanySettings", "Required", "Complete", "Medium", "Update settings by company"],

    # ==================== SYSTEM - NOTIFICATION PREFERENCES APIs ====================
    ["System - Notifications", "/api/notification-preferences/mine", "GET", "getMyNotificationPreferences", "Required", "Complete", "Medium", "Get my preferences"],
    ["System - Notifications", "/api/notification-preferences/mine", "PUT", "upsertMyNotificationPreferences", "Required", "Complete", "Medium", "Update my preferences"],
    ["System - Notifications", "/api/notification-preferences/user/:userId", "GET", "getUserNotificationPreferences", "Required", "Complete", "Medium", "Get user preferences"],

    # ==================== SYSTEM - TABLE CRUD APIs ====================
    ["System - Table CRUD", "/api/tables/", "GET", "listTables", "Required", "Complete", "Medium", "List all tables"],
    ["System - Table CRUD", "/api/tables/:tableName/meta", "GET", "getTableMeta", "Required", "Complete", "Medium", "Get table metadata"],
    ["System - Table CRUD", "/api/tables/:tableName/rows", "GET", "listRows", "Required", "Complete", "Medium", "List table rows"],
    ["System - Table CRUD", "/api/tables/:tableName/rows/:id", "GET", "getRowById", "Required", "Complete", "Medium", "Get row by ID"],
    ["System - Table CRUD", "/api/tables/:tableName/rows", "POST", "createRow", "Required", "Complete", "Medium", "Create row"],
    ["System - Table CRUD", "/api/tables/:tableName/rows/:id", "PUT", "updateRow", "Required", "Complete", "Medium", "Update row"],
    ["System - Table CRUD", "/api/tables/:tableName/rows/:id", "DELETE", "deleteRow", "Required", "Complete", "Medium", "Delete row"],

    # ==================== SYSTEM - TEAMS CHAT APIs ====================
    ["System - Teams Chat", "/api/teams-chat/users", "GET", "getCompanyUsers", "Required", "Complete", "Medium", "Get company users"],
    ["System - Teams Chat", "/api/teams-chat/teams", "GET", "getTeams", "Required", "Complete", "Medium", "Get all teams"],
    ["System - Teams Chat", "/api/teams-chat/teams", "POST", "createTeam", "Required", "Complete", "Medium", "Create team"],
    ["System - Teams Chat", "/api/teams-chat/teams/:teamId/members", "GET", "getTeamMembers", "Required", "Complete", "Medium", "Get team members"],
    ["System - Teams Chat", "/api/teams-chat/teams/:teamId/members", "POST", "addTeamMembers", "Required", "Complete", "Medium", "Add team members"],
    ["System - Teams Chat", "/api/teams-chat/teams/:teamId/members/:userId", "DELETE", "removeTeamMember", "Required", "Complete", "Medium", "Remove team member"],
    ["System - Teams Chat", "/api/teams-chat/teams/:teamId/channels", "GET", "getTeamChannels", "Required", "Complete", "Medium", "Get team channels"],
    ["System - Teams Chat", "/api/teams-chat/teams/:teamId/channels", "POST", "createChannel", "Required", "Complete", "Medium", "Create channel"],
    ["System - Teams Chat", "/api/teams-chat/channels/:channelId/messages", "GET", "getChannelMessages", "Required", "Complete", "Medium", "Get channel messages"],
    ["System - Teams Chat", "/api/teams-chat/channels/:channelId/messages", "POST", "sendMessage", "Required", "Complete", "Medium", "Send message"],
    ["System - Teams Chat", "/api/teams-chat/channels/:channelId/read", "POST", "markChannelAsRead", "Required", "Complete", "Medium", "Mark channel as read"],

    # ==================== USER TYPES APIs ====================
    ["User Types", "/api/user-types/", "GET", "getAllUserTypes", "Required", "Complete", "Medium", "Get all user types"],
    ["User Types", "/api/user-types/", "POST", "createUserType", "Required", "Complete", "Medium", "Create user type"],
    ["User Types", "/api/user-types/:id", "GET", "getUserTypeById", "Required", "Complete", "Medium", "Get user type by ID"],
    ["User Types", "/api/user-types/:id", "PUT", "updateUserType", "Required", "Complete", "Medium", "Update user type"],
    ["User Types", "/api/user-types/:id", "DELETE", "deleteUserType", "Required", "Complete", "Medium", "Delete user type"],

    # ==================== PURCHASE REQUISITION APIs ====================
    ["Purchase Requisitions", "/api/purchase-requisitions/", "GET", "getAllPurchaseRequisitions", "Required", "Complete", "High", "Get all PRs"],
    ["Purchase Requisitions", "/api/purchase-requisitions/", "POST", "createPurchaseRequisition", "Required", "Complete", "High", "Create PR"],
    ["Purchase Requisitions", "/api/purchase-requisitions/:id", "GET", "getPurchaseRequisitionById", "Required", "Complete", "High", "Get PR by ID"],
    ["Purchase Requisitions", "/api/purchase-requisitions/:id", "PUT", "updatePurchaseRequisition", "Required", "Complete", "High", "Update PR"],
    ["Purchase Requisitions", "/api/purchase-requisitions/:id", "DELETE", "deletePurchaseRequisition", "Required", "Complete", "High", "Delete PR"],

    # ==================== PURCHASE RETURN APIs ====================
    ["Purchase Returns", "/api/purchase-returns/", "GET", "getAllPurchaseReturns", "Required", "Complete", "High", "Get all returns"],
    ["Purchase Returns", "/api/purchase-returns/", "POST", "createPurchaseReturn", "Required", "Complete", "High", "Create return"],
    ["Purchase Returns", "/api/purchase-returns/:id", "GET", "getPurchaseReturnById", "Required", "Complete", "High", "Get return by ID"],
    ["Purchase Returns", "/api/purchase-returns/:id", "PUT", "updatePurchaseReturn", "Required", "Complete", "High", "Update return"],

    # ==================== SALES RETURN APIs ====================
    ["Sales Returns", "/api/sales-returns/", "GET", "getAllSalesReturns", "Required", "Complete", "High", "Get all returns"],
    ["Sales Returns", "/api/sales-returns/", "POST", "createSalesReturn", "Required", "Complete", "High", "Create return"],
    ["Sales Returns", "/api/sales-returns/:id", "GET", "getSalesReturnById", "Required", "Complete", "High", "Get return by ID"],
    ["Sales Returns", "/api/sales-returns/:id", "PUT", "updateSalesReturn", "Required", "Complete", "High", "Update return"],

    # ==================== DELIVERY CHALLAN APIs ====================
    ["Delivery Challans", "/api/delivery-challans/", "GET", "getAllDeliveryChallans", "Required", "Complete", "High", "Get all challans"],
    ["Delivery Challans", "/api/delivery-challans/", "POST", "createDeliveryChallan", "Required", "Complete", "High", "Create challan"],
    ["Delivery Challans", "/api/delivery-challans/:id", "GET", "getDeliveryChallanById", "Required", "Complete", "High", "Get challan by ID"],
    ["Delivery Challans", "/api/delivery-challans/:id", "PUT", "updateDeliveryChallan", "Required", "Complete", "High", "Update challan"],

    # ==================== SALES QUOTATION APIs ====================
    ["Sales Quotations", "/api/sales-quotations/", "GET", "getAllSalesQuotations", "Required", "Complete", "High", "Get all quotations"],
    ["Sales Quotations", "/api/sales-quotations/", "POST", "createSalesQuotation", "Required", "Complete", "High", "Create quotation"],
    ["Sales Quotations", "/api/sales-quotations/:id", "GET", "getSalesQuotationById", "Required", "Complete", "High", "Get quotation by ID"],
    ["Sales Quotations", "/api/sales-quotations/:id", "PUT", "updateSalesQuotation", "Required", "Complete", "High", "Update quotation"],

    # ==================== PRODUCTION APIs ====================
    ["Production - BOM", "/api/bom/", "GET", "getAllBOMs", "Required", "Complete", "High", "Get all BOMs"],
    ["Production - BOM", "/api/bom/", "POST", "createBOM", "Required", "Complete", "High", "Create BOM"],
    ["Production - BOM", "/api/bom/:id", "GET", "getBOMById", "Required", "Complete", "High", "Get BOM by ID"],
    ["Production - BOM", "/api/bom/:id", "PUT", "updateBOM", "Required", "Complete", "High", "Update BOM"],

    ["Production - Orders", "/api/production-orders/", "GET", "getAllProductionOrders", "Required", "Complete", "High", "Get all orders"],
    ["Production - Orders", "/api/production-orders/", "POST", "createProductionOrder", "Required", "Complete", "High", "Create order"],
    ["Production - Orders", "/api/production-orders/:id", "GET", "getProductionOrderById", "Required", "Complete", "High", "Get order by ID"],
    ["Production - Orders", "/api/production-orders/:id", "PUT", "updateProductionOrder", "Required", "Complete", "High", "Update order"],

    # ==================== APPROVAL APIs ====================
    ["Approvals", "/api/approvals/", "GET", "getAllApprovals", "Required", "Complete", "High", "Get all approvals"],
    ["Approvals", "/api/approvals/", "POST", "createApproval", "Required", "Complete", "High", "Create approval request"],
    ["Approvals", "/api/approvals/:id", "GET", "getApprovalById", "Required", "Complete", "High", "Get approval by ID"],
    ["Approvals", "/api/approvals/:id/approve", "POST", "approveRequest", "Required", "Complete", "High", "Approve request"],
    ["Approvals", "/api/approvals/:id/reject", "POST", "rejectRequest", "Required", "Complete", "High", "Reject request"],

    # ==================== EXPENSE APIs ====================
    ["Expenses", "/api/expenses/", "GET", "getAllExpenses", "Required", "Complete", "Medium", "Get all expenses"],
    ["Expenses", "/api/expenses/", "POST", "createExpense", "Required", "Complete", "Medium", "Create expense"],
    ["Expenses", "/api/expenses/:id", "GET", "getExpenseById", "Required", "Complete", "Medium", "Get expense by ID"],
    ["Expenses", "/api/expenses/:id", "PUT", "updateExpense", "Required", "Complete", "Medium", "Update expense"],
    ["Expenses", "/api/expenses/:id", "DELETE", "deleteExpense", "Required", "Complete", "Medium", "Delete expense"],

    # ==================== QUALITY CONTROL APIs ====================
    ["Quality Control", "/api/quality-control/", "GET", "getAllQualityControls", "Required", "Partial", "Medium", "Get all QC records"],
    ["Quality Control", "/api/quality-control/", "POST", "createQualityControl", "Required", "Partial", "Medium", "Create QC record"],
    ["Quality Control", "/api/quality-control/:id", "GET", "getQualityControlById", "Required", "Partial", "Medium", "Get QC by ID"],
]

# Write data
for row_idx, row_data in enumerate(api_data, 5):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws1.cell(row=row_idx, column=col_idx, value=value)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        
        # Color coding based on HTTP method
        if col_idx == 3:  # Method column
            if value == "GET":
                cell.fill = get_fill
            elif value == "POST":
                cell.fill = post_fill
            elif value == "PUT":
                cell.fill = put_fill
            elif value == "PATCH":
                cell.fill = patch_fill
            elif value == "DELETE":
                cell.fill = delete_fill

# Calculate statistics
total_apis = len(api_data)
complete_apis = sum(1 for row in api_data if row[6] == "Complete")
partial_apis = sum(1 for row in api_data if row[6] == "Partial")
not_started_apis = sum(1 for row in api_data if row[6] == "Not Started")

# Summary section
summary_row = total_apis + 7
ws1.merge_cells(f'A{summary_row}:H{summary_row}')
ws1[f'A{summary_row}'] = "API SUMMARY STATISTICS"
ws1[f'A{summary_row}'].font = Font(size=14, bold=True)
ws1[f'A{summary_row}'].alignment = Alignment(horizontal='center', vertical='center')

summary_data = [
    ["Total APIs", total_apis, "", "", "", "", "", ""],
    ["Complete", complete_apis, f"{round(complete_apis/total_apis*100, 1)}%", "", "", "", "", ""],
    ["Partial", partial_apis, f"{round(partial_apis/total_apis*100, 1)}%", "", "", "", "", ""],
    ["Not Started", not_started_apis, f"{round(not_started_apis/total_apis*100, 1)}%", "", "", "", "", ""],
]

for row_idx, row_data in enumerate(summary_data, summary_row + 1):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws1.cell(row=row_idx, column=col_idx, value=value)
        cell.border = thin_border
        if col_idx in [1, 2]:
            cell.font = Font(bold=True)

# Module-wise summary
module_summary_row = summary_row + 7
ws1.merge_cells(f'A{module_summary_row}:H{module_summary_row}')
ws1[f'A{module_summary_row}'] = "MODULE-WISE API COUNT"
ws1[f'A{module_summary_row}'].font = Font(size=14, bold=True)
ws1[f'A{module_summary_row}'].alignment = Alignment(horizontal='center', vertical='center')

# Calculate module-wise stats
modules = {}
for row in api_data:
    module = row[0]
    if module not in modules:
        modules[module] = {"total": 0, "complete": 0, "partial": 0}
    modules[module]["total"] += 1
    if row[6] == "Complete":
        modules[module]["complete"] += 1
    elif row[6] == "Partial":
        modules[module]["partial"] += 1

# Module summary headers
module_headers = ["Module", "Total APIs", "Complete", "Partial", "Not Started", "Completion %", "Status", ""]
for col, header in enumerate(module_headers, 1):
    cell = ws1.cell(row=module_summary_row + 1, column=col, value=header)
    cell.fill = header_fill
    cell.font = header_font
    cell.border = thin_border
    cell.alignment = Alignment(horizontal='center', vertical='center')

# Write module summary
row_num = module_summary_row + 2
for module, stats in modules.items():
    not_started = stats["total"] - stats["complete"] - stats["partial"]
    completion = round((stats["complete"] + (stats["partial"] * 0.5)) / stats["total"] * 100, 1)
    
    status = "Complete" if completion == 100 else "Partial" if completion > 0 else "Not Started"
    
    module_data = [
        module,
        stats["total"],
        stats["complete"],
        stats["partial"],
        not_started,
        f"{completion}%",
        status,
        ""
    ]
    
    for col_idx, value in enumerate(module_data, 1):
        cell = ws1.cell(row=row_num, column=col_idx, value=value)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        
        if col_idx == 7:  # Status column
            if value == "Complete":
                cell.fill = complete_fill
            elif value == "Partial":
                cell.fill = partial_fill
            elif value == "Not Started":
                cell.fill = not_started_fill
    
    row_num += 1

# Adjust column widths
column_widths = [25, 50, 12, 30, 15, 12, 12, 40]
for i, width in enumerate(column_widths, 1):
    ws1.column_dimensions[get_column_letter(i)].width = width

# Freeze panes
ws1.freeze_panes = 'A5'

# Save workbook
output_file = "COMPLETE_API_INVENTORY.xlsx"
wb.save(output_file)
print(f"✅ Excel file created: {output_file}")
print(f"\n📊 API Statistics:")
print(f"   Total APIs: {total_apis}")
print(f"   Complete: {complete_apis} ({round(complete_apis/total_apis*100, 1)}%)")
print(f"   Partial: {partial_apis} ({round(partial_apis/total_apis*100, 1)}%)")
print(f"   Not Started: {not_started_apis} ({round(not_started_apis/total_apis*100, 1)}%)")
print(f"\n📁 Modules Covered: {len(modules)}")
for module, stats in sorted(modules.items()):
    completion = round((stats["complete"] + (stats["partial"] * 0.5)) / stats["total"] * 100, 1)
    print(f"   • {module}: {stats['total']} APIs ({completion}%)")