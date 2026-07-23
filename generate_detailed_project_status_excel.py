import subprocess
from datetime import datetime

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


OUTPUT_FILE = "ERP_CRM_Project_Pending_Complete_Detailed.xlsx"


STATUS_COLORS = {
    "Complete": "C6EFCE",
    "Mostly Complete": "D9EAD3",
    "Partial": "FFF2CC",
    "Pending": "FCE4D6",
    "Not Started": "F4CCCC",
    "Needs Verification": "D9EAF7",
}


FEATURES = [
    ("ERP", "Master Data", "Item/Product Master", "Complete", 100, "High", "Phase 1", "Products backend/frontend CRUD exists."),
    ("ERP", "Master Data", "Product Category/Sub-category", "Complete", 100, "High", "Phase 1", "ProductCategories module exists."),
    ("ERP", "Master Data", "Units/UOM", "Complete", 100, "High", "Phase 1", "Units module exists; conversion depth still needs business validation."),
    ("ERP", "Master Data", "Warehouse/Location Master", "Complete", 100, "High", "Phase 1", "Warehouses module exists."),
    ("ERP", "Master Data", "Supplier/Vendor Master", "Complete", 100, "High", "Phase 1", "Suppliers module exists."),
    ("ERP", "Master Data", "Customer Master", "Complete", 100, "High", "Phase 1", "Customers module exists."),
    ("ERP", "Master Data", "Brand Master", "Complete", 100, "Medium", "Phase 1", "Brands module exists."),
    ("ERP", "Master Data", "Tax Configuration", "Complete", 100, "High", "Phase 1", "Taxes and ProductTaxMap modules exist."),
    ("ERP", "Master Data", "Batch/Serial Attributes", "Complete", 100, "High", "Phase 1", "BatchSerial tracking exists."),
    ("ERP", "Master Data", "Bin/Rack Master", "Partial", 50, "Medium", "Phase 2", "Basic warehouse locations exist; dedicated bin/rack workflow not confirmed."),
    ("ERP", "Master Data", "Barcode Setup/Printing", "Partial", 40, "Low", "Phase 3", "Barcode field/helpers exist; complete printing/scanning workflow not confirmed."),
    ("ERP", "Master Data", "HSN/SAC Code Master", "Needs Verification", 70, "Medium", "Phase 2", "New HSN model/controller/route exists; needs API/UI testing."),
    ("ERP", "Master Data", "Price List Master", "Needs Verification", 70, "Medium", "Phase 2", "New price list model/controller/route exists; needs API/UI testing."),
    ("ERP", "Master Data", "Size/Color Attributes", "Not Started", 0, "Low", "Phase 3", "No dedicated item variant attribute module found."),
    ("ERP", "Purchase", "Purchase Requisition", "Complete", 100, "High", "Phase 1", "PurchaseRequisitions module exists."),
    ("ERP", "Purchase", "Purchase Order", "Complete", 100, "High", "Phase 1", "PurchaseOrders with items exists."),
    ("ERP", "Purchase", "Goods Receipt Note", "Complete", 100, "High", "Phase 1", "GRN module exists."),
    ("ERP", "Purchase", "Purchase Return/Debit Note", "Complete", 100, "High", "Phase 1", "PurchaseReturns module exists."),
    ("ERP", "Purchase", "PO Approval Workflow", "Partial", 60, "High", "Phase 1", "Generic approval system exists; PO-specific flow not fully confirmed."),
    ("ERP", "Purchase", "Purchase Order Tracking", "Partial", 50, "Medium", "Phase 2", "Basic status tracking exists; advanced tracking not confirmed."),
    ("ERP", "Purchase", "Request for Quotation", "Needs Verification", 70, "Medium", "Phase 2", "New RFQ model/controller/route exists; needs full testing."),
    ("ERP", "Purchase", "PO-GRN-Invoice Matching", "Needs Verification", 70, "High", "Phase 2", "New invoice match model/controller/route exists; needs full testing."),
    ("ERP", "Purchase", "Blanket/Contract Orders", "Not Started", 0, "Low", "Phase 3", "No dedicated module found."),
    ("ERP", "Sales", "Sales Quotation", "Complete", 100, "High", "Phase 1", "SalesQuotations module exists."),
    ("ERP", "Sales", "Sales Order", "Complete", 100, "High", "Phase 1", "SalesOrders module exists."),
    ("ERP", "Sales", "Delivery Challan/Dispatch", "Complete", 100, "High", "Phase 1", "DeliveryChallans module exists."),
    ("ERP", "Sales", "Sales Return/Credit Note", "Complete", 100, "High", "Phase 1", "SalesReturns module exists."),
    ("ERP", "Sales", "Invoice Generation", "Complete", 100, "High", "Phase 1", "CRM invoices module exists."),
    ("ERP", "Sales", "Sales Order Approval", "Partial", 60, "High", "Phase 1", "Generic approval system exists; SO-specific flow not fully confirmed."),
    ("ERP", "Sales", "Backorder Management", "Partial", 40, "Medium", "Phase 2", "Basic handling may exist; dedicated tracking/reporting not found."),
    ("ERP", "Inventory", "Stock Inward Entry", "Complete", 100, "High", "Phase 1", "Covered through GRN process."),
    ("ERP", "Inventory", "Stock Outward Entry", "Complete", 100, "High", "Phase 1", "Covered through delivery challan/sales flow."),
    ("ERP", "Inventory", "Stock Transfer", "Complete", 100, "High", "Phase 1", "StockTransfers module exists."),
    ("ERP", "Inventory", "Stock Adjustment", "Complete", 100, "High", "Phase 1", "StockAdjustments module exists."),
    ("ERP", "Inventory", "Stock Movements/Ledger", "Complete", 100, "High", "Phase 1", "StockMovements module exists."),
    ("ERP", "Inventory", "Stock Reservation/Allocation", "Partial", 30, "Medium", "Phase 2", "Dedicated reservation module not found."),
    ("ERP", "Inventory", "Expiry Date Management", "Not Started", 0, "Medium", "Phase 2", "No complete expiry tracking workflow found."),
    ("ERP", "Inventory", "Stock Aging Report", "Not Started", 0, "Medium", "Phase 2", "No dedicated aging report found."),
    ("ERP", "Inventory", "Reorder Levels", "Needs Verification", 70, "High", "Phase 1", "New reorder model/controller/route and frontend page exist; needs end-to-end testing."),
    ("ERP", "Inventory", "Auto Replenishment", "Needs Verification", 50, "High", "Phase 1", "Reorder foundation exists; automatic PO/replenishment behavior needs validation."),
    ("ERP", "Inventory", "Min-Max Stock Settings", "Needs Verification", 60, "High", "Phase 1", "Likely part of new ReorderLevels; needs validation."),
    ("ERP", "WMS", "Warehouse Transfer Note", "Complete", 100, "High", "Phase 1", "Covered by StockTransfers."),
    ("ERP", "WMS", "Bin/Location Mapping", "Partial", 50, "Medium", "Phase 2", "Basic warehouse locations only."),
    ("ERP", "WMS", "Packing & Shipment", "Partial", 40, "Medium", "Phase 2", "Basic delivery challan workflow only."),
    ("ERP", "WMS", "Putaway Management", "Not Started", 0, "Medium", "Phase 3", "No dedicated workflow found."),
    ("ERP", "WMS", "Picking List Generation", "Not Started", 0, "Medium", "Phase 3", "No dedicated workflow found."),
    ("ERP", "WMS", "Cycle Count", "Not Started", 0, "Low", "Phase 3", "No dedicated workflow found."),
    ("ERP", "WMS", "Physical Stock Verification", "Not Started", 0, "Low", "Phase 3", "No dedicated workflow found."),
    ("ERP", "WMS", "Cross-Docking", "Not Started", 0, "Low", "Phase 4", "No dedicated workflow found."),
    ("ERP", "Costing", "Costing Method Setup", "Needs Verification", 70, "High", "Phase 1", "New costing method model and stock valuation files exist; needs testing."),
    ("ERP", "Costing", "Stock Valuation Report", "Needs Verification", 70, "High", "Phase 1", "New stock valuation route/controller/frontend page exists; needs testing."),
    ("ERP", "Costing", "Landed Cost Allocation", "Needs Verification", 55, "Medium", "Phase 2", "New landed cost model exists; route/UI completeness needs validation."),
    ("ERP", "Costing", "Cost Adjustment Entries", "Needs Verification", 55, "Medium", "Phase 2", "New cost adjustment model exists; route/UI completeness needs validation."),
    ("ERP", "Quality Control", "Quality Inspection Setup", "Partial", 30, "Medium", "Phase 2", "QualityControl model exists but integration is incomplete."),
    ("ERP", "Quality Control", "QC on GRN", "Not Started", 0, "Medium", "Phase 2", "No full GRN QC workflow confirmed."),
    ("ERP", "Quality Control", "Rejected/Quarantine Stock", "Not Started", 0, "Medium", "Phase 2", "No quarantine workflow confirmed."),
    ("ERP", "Quality Control", "QC Certificates", "Not Started", 0, "Low", "Phase 3", "No certificate workflow confirmed."),
    ("ERP", "Production", "Bill of Materials", "Complete", 100, "High", "Phase 1", "BOM module exists."),
    ("ERP", "Production", "Work Order/Job Card", "Complete", 100, "High", "Phase 1", "ProductionOrders module exists."),
    ("ERP", "Production", "Material Requisition", "Partial", 50, "Medium", "Phase 2", "May be part of production module; workflow needs validation."),
    ("ERP", "Production", "Raw Material Issue", "Partial", 50, "Medium", "Phase 2", "May be part of production module; workflow needs validation."),
    ("ERP", "Production", "Finished Goods Receipt", "Partial", 50, "Medium", "Phase 2", "May be part of production module; workflow needs validation."),
    ("ERP", "Production", "By-product/Scrap Entry", "Not Started", 0, "Low", "Phase 3", "No dedicated module found."),
    ("ERP", "Reports", "Stock Movement Report", "Complete", 100, "High", "Phase 1", "StockMovements module exists."),
    ("ERP", "Reports", "Batch/Serial Traceability", "Complete", 100, "High", "Phase 1", "BatchSerial module exists."),
    ("ERP", "Reports", "Profit/Loss Reports", "Complete", 100, "High", "Phase 1", "ProfitLossReports module exists."),
    ("ERP", "Reports", "Stock Summary", "Partial", 50, "Medium", "Phase 2", "Basic stock reporting exists."),
    ("ERP", "Reports", "Item-wise Sales/Purchase Analysis", "Partial", 60, "Medium", "Phase 2", "Basic reporting exists; advanced analytics not confirmed."),
    ("ERP", "Reports", "ABC Analysis", "Not Started", 0, "Medium", "Phase 3", "No dedicated analytics found."),
    ("ERP", "Reports", "Slow/Non-moving Stock", "Not Started", 0, "Medium", "Phase 3", "No dedicated report found."),
    ("ERP", "Reports", "Vendor Performance", "Not Started", 0, "Medium", "Phase 3", "No dedicated report found."),
    ("ERP", "Reports", "GRN vs PO Reconciliation", "Not Started", 0, "Medium", "Phase 2", "No dedicated reconciliation report confirmed."),
    ("ERP", "Utilities", "Import/Export Excel/CSV", "Complete", 100, "High", "Phase 1", "DataImportExport module exists."),
    ("ERP", "Utilities", "Approval Workflow Configuration", "Complete", 100, "High", "Phase 1", "Approval workflow tables exist."),
    ("ERP", "Utilities", "Audit Trail/Logs", "Complete", 100, "High", "Phase 1", "AuditLogs and security tracking exist."),
    ("ERP", "Utilities", "Multi-Currency", "Not Started", 0, "Medium", "Phase 2", "No multi-currency support found."),
    ("ERP", "Utilities", "Financial Year/Periods", "Needs Verification", 70, "Medium", "Phase 2", "New financial year model/controller/route exists; needs testing."),
    ("CRM", "Core CRM", "Leads Management", "Complete", 100, "High", "Phase 1", "Full lead entity with source, status, rating, assignment, conversion fields."),
    ("CRM", "Core CRM", "Opportunities Management", "Complete", 100, "High", "Phase 1", "Opportunity budget, stage, probability, close handling exists."),
    ("CRM", "Core CRM", "Accounts Management", "Complete", 100, "High", "Phase 1", "Account CRUD and ownership fields exist."),
    ("CRM", "Core CRM", "Contacts Management", "Complete", 100, "High", "Phase 1", "Contact details and account association exist."),
    ("CRM", "Core CRM", "Activities Management", "Complete", 100, "High", "Phase 1", "Activities with type, due dates, status, priority, reminders."),
    ("CRM", "Core CRM", "Cases/Support Tickets", "Complete", 100, "High", "Phase 1", "Cases with status, priority, assignment, resolution."),
    ("CRM", "Core CRM", "Quotes Management", "Complete", 100, "High", "Phase 1", "Quote numbering, totals, status, terms."),
    ("CRM", "Core CRM", "Invoices Management", "Complete", 100, "High", "Phase 1", "Invoice numbering, payment status, due date, totals."),
    ("CRM", "Core CRM", "Payments Management", "Complete", 100, "High", "Phase 1", "Payment recording, methods, references, invoice association."),
    ("CRM", "Core CRM", "Retentions/Follow-ups", "Complete", 100, "Medium", "Phase 1", "Retention type, status, next action, reminders."),
    ("CRM", "Core CRM", "Presales Management", "Partial", 70, "Medium", "Phase 2", "Model and page exist; integration/details need verification."),
    ("CRM", "Master Data", "Lead Sources", "Complete", 100, "High", "Phase 1", "Master data exists."),
    ("CRM", "Master Data", "Industries", "Complete", 100, "Medium", "Phase 1", "Master data exists."),
    ("CRM", "Master Data", "Sales Stages", "Complete", 100, "High", "Phase 1", "Master data exists."),
    ("CRM", "Master Data", "Follow-up Types", "Complete", 100, "Medium", "Phase 1", "Master data exists."),
    ("CRM", "Master Data", "Task Types", "Complete", 100, "Medium", "Phase 1", "Master data exists."),
    ("CRM", "Master Data", "Payment Terms", "Not Started", 0, "Medium", "Phase 2", "No dedicated master found."),
    ("CRM", "Master Data", "Case Categories/Resolution Codes", "Not Started", 0, "Low", "Phase 3", "No dedicated master found."),
    ("CRM", "Workflow", "Lead to Opportunity", "Complete", 100, "High", "Phase 1", "Conversion workflow exists."),
    ("CRM", "Workflow", "Opportunity to Quote", "Complete", 100, "High", "Phase 1", "Conversion workflow exists."),
    ("CRM", "Workflow", "Quote to Invoice", "Complete", 100, "High", "Phase 1", "Conversion workflow exists."),
    ("CRM", "Workflow", "Invoice to Payment", "Complete", 100, "High", "Phase 1", "Payment tracking exists."),
    ("CRM", "Workflow", "Automations and Reminders", "Partial", 60, "Medium", "Phase 2", "CRM jobs and automation services exist; coverage needs validation."),
    ("CRM", "Workflow", "Visual Workflow Builder", "Not Started", 0, "Medium", "Phase 3", "No builder found."),
    ("CRM", "Workflow", "SLA Management", "Partial", 40, "Medium", "Phase 3", "Case SLA job exists; full configuration flow not confirmed."),
    ("CRM", "Analytics", "Sales Pipeline Dashboard", "Complete", 100, "High", "Phase 1", "CRM dashboard/components exist."),
    ("CRM", "Analytics", "Lead/Opportunity Reports", "Complete", 100, "High", "Phase 1", "Basic reporting exists."),
    ("CRM", "Analytics", "Conversion Analytics", "Not Started", 0, "Medium", "Phase 2", "Dedicated analytics not confirmed."),
    ("CRM", "Analytics", "Sales Forecasting", "Not Started", 0, "Medium", "Phase 2", "No dedicated forecasting module found."),
    ("CRM", "Analytics", "Custom Report Builder", "Not Started", 0, "Low", "Phase 3", "No builder found."),
    ("CRM", "Integration", "Email Sending/Templates/Tracking", "Pending", 20, "High", "Phase 2", "Nodemailer exists; full CRM email integration not confirmed."),
    ("CRM", "Integration", "Calendar Sync", "Not Started", 0, "Medium", "Phase 2", "No calendar integration found."),
    ("CRM", "Integration", "Document Management", "Needs Verification", 70, "Medium", "Phase 2", "New document model/controller/route exists; needs CRM entity integration testing."),
    ("CRM", "Integration", "Customer Portal", "Partial", 40, "Low", "Phase 3", "UserPortal components exist; customer portal completeness not confirmed."),
    ("CRM", "Integration", "SMS/Social/Call Recording", "Not Started", 0, "Low", "Phase 4", "No integrations found."),
    ("User/RBAC", "Authentication", "JWT Login/Logout/Refresh", "Complete", 100, "High", "Phase 1", "Auth service, refresh tokens, session management exist."),
    ("User/RBAC", "Authentication", "Password Reset/Email Verification", "Complete", 100, "High", "Phase 1", "Services and routes documented as complete."),
    ("User/RBAC", "Authentication", "2FA/TOTP", "Needs Verification", 70, "High", "Phase 1", "New User2FA model/controller/routes exist; needs login-flow testing."),
    ("User/RBAC", "Authentication", "OAuth", "Not Started", 0, "Medium", "Phase 3", "Explicitly listed as not implemented."),
    ("User/RBAC", "Authentication", "LDAP/Active Directory", "Not Started", 0, "Medium", "Phase 3", "Explicitly listed as not implemented."),
    ("User/RBAC", "Security", "Login History/Suspicious Detection", "Complete", 100, "High", "Phase 1", "LoginHistory and security services exist."),
    ("User/RBAC", "Security", "IP Whitelisting", "Not Started", 0, "Medium", "Phase 3", "Explicitly listed as not implemented."),
    ("User/RBAC", "RBAC", "Roles and Permissions", "Complete", 100, "High", "Phase 1", "Database-driven RBAC with modules/permissions/menus."),
    ("User/RBAC", "RBAC", "Menu Permissions", "Complete", 100, "High", "Phase 1", "Menu visibility control exists."),
    ("User/RBAC", "RBAC", "Company Isolation", "Complete", 100, "High", "Phase 1", "Company scope middleware/utilities exist."),
    ("User/RBAC", "RBAC", "User Type Management UI", "Needs Verification", 70, "High", "Phase 1", "New frontend UserTypes page exists; needs testing."),
    ("User/RBAC", "Users", "User Registration/Profile/List", "Complete", 100, "High", "Phase 1", "Frontend and backend pages/controllers exist."),
    ("User/RBAC", "Users", "Organization Chart", "Complete", 100, "Medium", "Phase 1", "Classic org chart component exists."),
    ("User/RBAC", "Users", "Bulk User Import", "Partial", 40, "Medium", "Phase 2", "Import utilities exist; user-specific flow not confirmed."),
    ("User/RBAC", "Users", "User Impersonation", "Not Started", 0, "Medium", "Phase 3", "No implementation found."),
]


NEW_VERIFICATION_ITEMS = [
    ("Inventory", "Stock Valuation", "ERPCRMServer/Models/InventoryManagement/StockValuation.js; routes/Inventory/stockValuation", "Backend and frontend page appear newly added.", "Run API smoke test and confirm valuation formulas."),
    ("Inventory", "Costing Method", "ERPCRMServer/Models/InventoryManagement/CostingMethod.js", "Model exists.", "Confirm controller coverage and DB table creation."),
    ("Inventory", "Landed Cost", "ERPCRMServer/Models/InventoryManagement/LandedCost.js", "Model exists.", "Confirm route/controller/UI coverage."),
    ("Inventory", "Cost Adjustment", "ERPCRMServer/Models/InventoryManagement/CostAdjustment.js", "Model exists.", "Confirm route/controller/UI coverage."),
    ("Inventory", "Reorder Levels", "ERPCRMServer/controllers/InventoryApis/reorderLevelController.js; clientui/src/features/inventory/pages/ReorderLevelsPage.jsx", "Backend route and frontend page exist.", "Test create/update alerts and auto replenish."),
    ("Inventory", "RFQ", "ERPCRMServer/controllers/InventoryApis/rfqController.js; routes/Inventory/rfq", "Backend files exist.", "Test RFQ lifecycle and supplier linking."),
    ("Inventory", "Invoice Matching", "ERPCRMServer/controllers/InventoryApis/invoiceMatchController.js; routes/Inventory/invoiceMatch", "Backend files exist.", "Test PO-GRN-Invoice 3-way matching."),
    ("Inventory", "HSN Code", "ERPCRMServer/controllers/InventoryApis/hsnCodeController.js; routes/Inventory/hsnCode", "Backend files exist.", "Test CRUD and product/tax integration."),
    ("Inventory", "Price List", "ERPCRMServer/controllers/InventoryApis/priceListController.js; routes/Inventory/priceList", "Backend files exist.", "Test pricing rules and frontend access."),
    ("Inventory", "Financial Year", "ERPCRMServer/controllers/InventoryApis/financialYearController.js; routes/Inventory/financialYear", "Backend files exist.", "Test active year, period close, accounting rules."),
    ("CRM/System", "Documents", "ERPCRMServer/controllers/InventoryApis/documentController.js; routes/Inventory/documents", "Document management backend exists.", "Confirm CRM entity attachments/versioning/share flow."),
    ("Auth", "Two Factor Authentication", "ERPCRMServer/controllers/auth/twoFactorController.js; routes/Auth/twoFactorRoutes.js", "2FA backend files exist.", "Test setup, verify, disable, backup codes, login enforcement."),
    ("Database", "Schema Conflict Fix", "ERPCRMServer/docs/SCHEMA_FIX_REPORT.md; scripts/fix_database_schema.js", "Fix documented for ModuleId/PermissionId conflicts.", "Run verify_schema.js against target DB."),
]


ROADMAP = [
    ("Phase 1", "Critical Core", "Months 1-4", "Stock valuation, reorder levels, 2FA, documents, audit improvements", "Raise overall completion to about 75%"),
    ("Phase 2", "Process Enhancement", "Months 5-8", "QC, financial settings, RFQ/invoice matching, reports", "Raise overall completion to about 85%"),
    ("Phase 3", "Advanced Features", "Months 9-12", "WMS, analytics, enterprise user/security features", "Raise overall completion to about 95%"),
    ("Phase 4", "Integrations & Polish", "Months 13-16", "E-commerce, logistics, calendar/SMS/social integrations, final hardening", "Reach 100%"),
]


def git_lines(args):
    try:
        result = subprocess.run(["git", *args], capture_output=True, text=True, check=False)
        return [line for line in result.stdout.splitlines() if line.strip()]
    except Exception:
        return []


def style_sheet(ws):
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border
            if cell.row == 1:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor="1F4E78")
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for col in range(1, ws.max_column + 1):
        letter = get_column_letter(col)
        width = min(max(len(str(ws.cell(row=r, column=col).value or "")) for r in range(1, ws.max_row + 1)) + 2, 55)
        ws.column_dimensions[letter].width = max(width, 12)


def apply_status_colors(ws, status_col):
    for row in range(2, ws.max_row + 1):
        status = ws.cell(row=row, column=status_col).value
        color = STATUS_COLORS.get(status)
        if color:
            for col in range(1, ws.max_column + 1):
                ws.cell(row=row, column=col).fill = PatternFill("solid", fgColor=color)


def add_table(ws, headers, rows):
    ws.append(headers)
    for row in rows:
        ws.append(row)
    style_sheet(ws)


def main():
    wb = Workbook()
    wb.remove(wb.active)

    generated_at = datetime.now().strftime("%Y-%m-%d %H:%M")
    counts = {}
    for _, _, _, status, *_ in FEATURES:
        counts[status] = counts.get(status, 0) + 1
    total = len(FEATURES)
    avg_completion = round(sum(row[4] for row in FEATURES) / total, 1)

    ws = wb.create_sheet("Executive Summary")
    summary_rows = [
        ("Report Generated", generated_at),
        ("Project", "ERP/CRM Full Project"),
        ("Source", "Local repo docs, route/model/page scan, and git status"),
        ("Overall completion from roadmap", "52.9% complete / 47.1% remaining"),
        ("Detailed feature rows in this workbook", total),
        ("Average completion from detailed rows", f"{avg_completion}%"),
        ("Complete", counts.get("Complete", 0)),
        ("Mostly Complete", counts.get("Mostly Complete", 0)),
        ("Partial", counts.get("Partial", 0)),
        ("Needs Verification", counts.get("Needs Verification", 0)),
        ("Pending", counts.get("Pending", 0)),
        ("Not Started", counts.get("Not Started", 0)),
        ("Main strength", "CRM core flow and RBAC/auth foundation are strong."),
        ("Main risk", "Several newly added ERP/Auth features need database/API/frontend verification before calling them complete."),
    ]
    add_table(ws, ["Metric", "Value"], summary_rows)

    ws = wb.create_sheet("All Feature Details")
    add_table(ws, ["Area", "Module", "Feature", "Status", "Completion %", "Priority", "Phase", "Details / Evidence"], FEATURES)
    apply_status_colors(ws, 4)

    ws = wb.create_sheet("Completed")
    completed = [row for row in FEATURES if row[3] in ("Complete", "Mostly Complete")]
    add_table(ws, ["Area", "Module", "Feature", "Status", "Completion %", "Priority", "Phase", "Details / Evidence"], completed)
    apply_status_colors(ws, 4)

    ws = wb.create_sheet("Pending Partial")
    pending_partial = [row for row in FEATURES if row[3] in ("Partial", "Pending")]
    add_table(ws, ["Area", "Module", "Feature", "Status", "Completion %", "Priority", "Phase", "Details / Evidence"], pending_partial)
    apply_status_colors(ws, 4)

    ws = wb.create_sheet("Not Started")
    not_started = [row for row in FEATURES if row[3] == "Not Started"]
    add_table(ws, ["Area", "Module", "Feature", "Status", "Completion %", "Priority", "Phase", "Details / Evidence"], not_started)
    apply_status_colors(ws, 4)

    ws = wb.create_sheet("Needs Verification")
    needs_verification = [row for row in FEATURES if row[3] == "Needs Verification"]
    add_table(ws, ["Area", "Module", "Feature", "Status", "Completion %", "Priority", "Phase", "Details / Evidence"], needs_verification)
    apply_status_colors(ws, 4)

    ws = wb.create_sheet("New Files To Verify")
    add_table(ws, ["Area", "Feature", "Files / Evidence", "Current Finding", "Next Validation Step"], NEW_VERIFICATION_ITEMS)

    ws = wb.create_sheet("Roadmap")
    add_table(ws, ["Phase", "Focus", "Timeline", "Major Work", "Target"], ROADMAP)

    status_lines = git_lines(["status", "--short"])
    ws = wb.create_sheet("Git Pending Changes")
    git_rows = []
    for line in status_lines:
        code = line[:2].strip() or "?"
        path = line[3:] if len(line) > 3 else line
        if code == "M":
            change_type = "Modified"
        elif code == "D":
            change_type = "Deleted"
        elif code == "??":
            change_type = "Untracked/New"
        else:
            change_type = code
        git_rows.append((change_type, path))
    add_table(ws, ["Change Type", "Path"], git_rows or [("Clean/Unavailable", "No git status rows captured")])

    ws = wb.create_sheet("Priority Next Actions")
    next_actions = [
        ("P0", "Run backend schema verification", "ERPCRMServer/scripts/verify_schema.js", "Confirms database fix and new tables."),
        ("P0", "Smoke test backend APIs", "npm test or route-level Postman collection", "Confirms new routes do not break startup/API behavior."),
        ("P1", "Validate stock valuation", "Stock valuation page and API", "High business impact; must verify formulas."),
        ("P1", "Validate reorder levels", "Reorder levels page and API", "High operations impact; confirm min/max and alert logic."),
        ("P1", "Validate 2FA login flow", "Auth 2FA routes", "Security feature must be tested end to end."),
        ("P1", "Validate RFQ and invoice matching", "RFQ/invoiceMatch routes", "Important purchase workflow gap."),
        ("P2", "Complete missing WMS workflows", "Putaway, picking, cycle count", "Needed for advanced warehouse operations."),
        ("P2", "Complete reporting gaps", "ABC, aging, vendor performance, reconciliation", "Needed for management visibility."),
        ("P3", "Add external integrations", "Email/calendar/SMS/e-commerce/logistics", "Useful after core workflows are stable."),
    ]
    add_table(ws, ["Priority", "Action", "Scope", "Reason"], next_actions)

    wb.save(OUTPUT_FILE)
    print(OUTPUT_FILE)


if __name__ == "__main__":
    main()
