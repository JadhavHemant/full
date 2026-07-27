
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

wb = openpyxl.Workbook()
ws1 = wb.active
ws1.title = "Feature Status"

# Title
ws1['A1'] = "ERP/CRM Project - Done & Pending Work Tracker"
ws1['A1'].font = Font(size=16, bold=True, color="FFFFFF")
ws1['A1'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
ws1['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws1.merge_cells('A1:H1')
ws1.row_dimensions[1].height = 30

# Subtitle
for col, val in enumerate(["Generated: 2026-07-24", "Total: 174", "Completed: 82", "Partial: 20", "Not Started: 72", "Completion: 52.9%"], 1):
    ws1.cell(row=2, column=col, value=val).font = Font(bold=True, size=10)

# Headers
headers = ['#', 'Module', 'Feature', 'Status', 'Completion %', 'Priority', 'Category', 'Notes']
for col, header in enumerate(headers, 1):
    cell = ws1.cell(row=4, column=col, value=header)
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

# Column widths
for col, w in enumerate([6, 30, 35, 14, 14, 10, 18, 40], 1):
    ws1.column_dimensions[openpyxl.utils.cell.get_column_letter(col)].width = w

features = [
    [1, "1. Master Data Management", "Item/Product Master", "COMPLETED", 100, "High", "Done", "Full CRUD with all fields"],
    [2, "1. Master Data Management", "Item Category / Sub-Category", "COMPLETED", 100, "High", "Done", "ProductCategories module"],
    [3, "1. Master Data Management", "Unit of Measurement (UOM)", "COMPLETED", 100, "High", "Done", "Units module with CRUD"],
    [4, "1. Master Data Management", "Warehouse/Location Master", "COMPLETED", 100, "High", "Done", "Warehouses module"],
    [5, "1. Master Data Management", "Supplier/Vendor Master", "COMPLETED", 100, "High", "Done", "Suppliers module"],
    [6, "1. Master Data Management", "Customer Master", "COMPLETED", 100, "High", "Done", "Customers module"],
    [7, "1. Master Data Management", "Brand Master", "COMPLETED", 100, "Medium", "Done", "Brands module"],
    [8, "1. Master Data Management", "Item Attributes (Batch, Serial)", "COMPLETED", 100, "High", "Done", "BatchSerial module"],
    [9, "1. Master Data Management", "Tax Configuration", "COMPLETED", 100, "High", "Done", "Taxes and ProductTaxMap"],
    [10, "1. Master Data Management", "Bin/Rack Master", "PARTIAL", 50, "Medium", "Partial-Complete", "Basic locations exist"],
    [11, "1. Master Data Management", "Item Barcode Setup", "PARTIAL", 40, "Low", "Partial-Complete", "Barcode field exists"],
    [12, "1. Master Data Management", "HSN/SAC Code Master", "NOT STARTED", 0, "Medium", "Pending", "Not implemented"],
    [13, "1. Master Data Management", "Item Attributes (Size, Color)", "NOT STARTED", 0, "Low", "Pending", "Only Batch/Serial tracked"],
    [14, "1. Master Data Management", "Price List Master", "NOT STARTED", 0, "Medium", "Pending", "Not implemented"],
    [15, "2. Purchase Management", "Purchase Requisition (PR)", "COMPLETED", 100, "High", "Done", "PurchaseRequisitions module"],
    [16, "2. Purchase Management", "Purchase Order (PO)", "COMPLETED", 100, "High", "Done", "PurchaseOrders with items"],
    [17, "2. Purchase Management", "Goods Receipt Note (GRN)", "COMPLETED", 100, "High", "Done", "GRN module"],
    [18, "2. Purchase Management", "Purchase Return / Debit Note", "COMPLETED", 100, "High", "Done", "PurchaseReturns module"],
    [19, "2. Purchase Management", "Purchase Order Approval", "PARTIAL", 60, "High", "Partial-Complete", "Generic approval system"],
    [20, "2. Purchase Management", "Purchase Order Tracking", "PARTIAL", 50, "Medium", "Partial-Complete", "Basic status tracking"],
    [21, "2. Purchase Management", "Request for Quotation (RFQ)", "NOT STARTED", 0, "Medium", "Pending", "Not implemented"],
    [22, "2. Purchase Management", "Vendor Bill/Invoice Matching", "NOT STARTED", 0, "High", "Pending", "No 3-way matching"],
    [23, "2. Purchase Management", "Blanket/Contract Orders", "NOT STARTED", 0, "Low", "Pending", "Not implemented"],
    [24, "3. Sales & Order Management", "Sales Quotation", "COMPLETED", 100, "High", "Done", "SalesQuotations module"],
    [25, "3. Sales & Order Management", "Sales Order (SO)", "COMPLETED", 100, "High", "Done", "SalesOrders module"],
    [26, "3. Sales & Order Management", "Sales Order Approval", "PARTIAL", 60, "High", "Partial-Complete", "Generic approval system"],
    [27, "3. Sales & Order Management", "Delivery Challan / Dispatch Note", "COMPLETED", 100, "High", "Done", "DeliveryChallans module"],
    [28, "3. Sales & Order Management", "Sales Return / Credit Note", "COMPLETED", 100, "High", "Done", "SalesReturns module"],
    [29, "3. Sales & Order Management", "Invoice Generation", "COMPLETED", 100, "High", "Done", "Invoices module (CRM)"],
    [30, "3. Sales & Order Management", "Backorder Management", "PARTIAL", 40, "Medium", "Partial-Complete", "Basic handling only"],
    [31, "4. Stock/Inventory Operations", "Stock Inward Entry", "COMPLETED", 100, "High", "Done", "Part of GRN process"],
    [32, "4. Stock/Inventory Operations", "Stock Outward Entry", "COMPLETED", 100, "High", "Done", "Part of DeliveryChallan"],
    [33, "4. Stock/Inventory Operations", "Stock Transfer (Inter-warehouse)", "COMPLETED", 100, "High", "Done", "StockTransfers module"],
    [34, "4. Stock/Inventory Operations", "Stock Adjustment (Positive/Negative)", "COMPLETED", 100, "High", "Done", "StockAdjustments module"],
    [35, "4. Stock/Inventory Operations", "Batch/Lot Tracking", "COMPLETED", 100, "High", "Done", "BatchSerial module"],
    [36, "4. Stock/Inventory Operations", "Serial Number Tracking", "COMPLETED", 100, "High", "Done", "BatchSerial module"],
    [37, "4. Stock/Inventory Operations", "Stock Movements", "COMPLETED", 100, "High", "Done", "StockMovements module"],
    [38, "4. Stock/Inventory Operations", "Stock Reservation/Allocation", "PARTIAL", 30, "Medium", "Partial-Complete", "Basic allocation only"],
    [39, "4. Stock/Inventory Operations", "Expiry Date Management", "NOT STARTED", 0, "Medium", "Pending", "Not implemented"],
    [40, "4. Stock/Inventory Operations", "Stock Aging Report", "NOT STARTED", 0, "Medium", "Pending", "Not implemented"],
    [41, "4. Stock/Inventory Operations", "Reorder Level & Auto Replenishment", "NOT STARTED", 0, "High", "Pending", "Not implemented"],
    [42, "4. Stock/Inventory Operations", "Min-Max Stock Settings", "NOT STARTED", 0, "High", "Pending", "Not implemented"],
    [43, "5. Warehouse Management (WMS)", "Bin/Location Mapping", "PARTIAL", 50, "Medium", "Partial-Complete", "Basic warehouse locations"],
    [44, "5. Warehouse Management (WMS)", "Warehouse Transfer Note", "COMPLETED", 100, "High", "Done", "StockTransfers module"],
    [45, "5. Warehouse Management (WMS)", "Packing & Shipment", "PARTIAL", 40, "Medium", "Partial-Complete", "Basic delivery challan"],
    [46, "5. Warehouse Management (WMS)", "Putaway Management", "NOT STARTED", 0, "Low", "Pending", "Not implemented"],
    [47, "5. Warehouse Management (WMS)", "Picking List Generation", "NOT STARTED", 0, "Low", "Pending", "Not implemented"],
    [48, "5. Warehouse Management (WMS)", "Cycle Count", "NOT STARTED", 0, "Low", "Pending", "Not implemented"],
    [49, "5. Warehouse Management (WMS)", "Physical Stock Verification", "NOT STARTED", 0, "Low", "Pending", "Not implemented"],
    [50, "5. Warehouse Management (WMS)", "Cross-Docking", "NOT STARTED", 0, "Low", "Pending", "Not implemented"],
    [51, "6. Stock Valuation & Costing", "Costing Method Setup", "NOT STARTED", 0, "High", "Pending", "Not implemented"],
    [52, "6. Stock Valuation & Costing", "Stock Valuation Report", "NOT STARTED", 0, "High", "Pending", "Not implemented"],
    [53, "6. Stock Valuation & Costing", "Landed Cost Allocation", "NOT STARTED", 0, "Medium", "Pending", "Not implemented"],
    [54, "6. Stock Valuation & Costing", "Cost Adjustment Entries", "NOT STARTED", 0, "Medium", "Pending", "Not implemented"],
    [55, "7. Quality Control", "Quality Inspection Setup", "PARTIAL", 30, "Medium", "Partial-Complete", "Model exists, not integrated"],
    [56, "7. Quality Control", "Incoming Inspection (QC on GRN)", "NOT STARTED", 0, "Medium", "Pending", "Not integrated with GRN"],
    [57, "7. Quality Control", "Rejected/Quarantine Stock", "NOT STARTED", 0, "Medium", "Pending", "Not implemented"],
    [58, "7. Quality Control", "QC Certificate Management", "NOT STARTED", 0, "Low", "Pending", "Not implemented"],
    [59, "8. Production/Manufacturing", "Bill of Materials (BOM)", "COMPLETED", 100, "High", "Done", "BOM module"],
    [60, "8. Production/Manufacturing", "Work Order / Job Card", "COMPLETED", 100, "High", "Done", "ProductionOrders module"],
    [61, "8. Production/Manufacturing", "Material Requisition for Production", "PARTIAL", 50, "Medium", "Partial-Complete", "Part of production module"],
    [62, "8. Production/Manufacturing", "Raw Material Issue", "PARTIAL", 50, "Medium", "Partial-Complete", "Part of production module"],
    [63, "8. Production/Manufacturing", "Finished Goods Receipt", "PARTIAL", 50, "Medium", "Partial-Complete", "Part of production module"],
    [64, "8. Production/Manufacturing", "By-product/Scrap Entry", "NOT STARTED", 0, "Low", "Pending", "Not implemented"],
    [65, "9. Reports & Analytics", "Stock Ledger / Movement Report", "COMPLETED", 100, "High", "Done", "StockMovements module"],
    [66, "9. Reports & Analytics", "Batch/Serial Traceability Report", "COMPLETED", 100, "High", "Done", "BatchSerial module"],
    [67, "9. Reports & Analytics", "Profit/Loss Reports", "COMPLETED", 100, "High", "Done", "ProfitLossReports module"],
    [68, "9. Reports & Analytics", "Stock Summary Report", "PARTIAL", 50, "Medium", "Partial-Complete", "Basic reports exist"],
    [69, "9. Reports & Analytics", "Item-wise Sales & Purchase Analysis", "PARTIAL", 60, "Medium", "Partial-Complete", "Basic reporting exists"],
    [70, "9. Reports & Analytics", "Stock Valuation Report", "NOT STARTED", 0, "High", "Pending", "Not implemented"],
    [71, "9. Reports & Analytics", "ABC Analysis", "NOT STARTED", 0, "Medium", "Pending", "Not implemented"],
    [72, "9. Reports & Analytics", "Slow-Moving/Non-Moving Stock Report", "NOT STARTED", 0, "Medium", "Pending", "Not implemented"],
    [73, "9. Reports & Analytics", "Reorder Report", "NOT STARTED", 0, "High", "Pending", "Not implemented"],
    [74, "9. Reports & Analytics", "Vendor Performance Report", "NOT STARTED", 0, "Medium", "Pending", "Not implemented"],
    [75, "9. Reports & Analytics", "Stock Aging Report", "NOT STARTED", 0, "Medium", "Pending", "Not implemented"],
    [76, "9. Reports & Analytics", "GRN vs PO Reconciliation", "NOT STARTED", 0, "Medium", "Pending", "Not implemented"],
    [77, "10. Utilities & Settings", "Import/Export Data (Excel/CSV)", "COMPLETED", 100, "High", "Done", "DataImportExport module"],
    [78, "10. Utilities & Settings", "User Role & Permission Setup", "COMPLETED", 100, "High", "Done", "Full RBAC system"],
    [79, "10. Utilities & Settings", "Approval Workflow Configuration", "COMPLETED", 100, "High", "Done", "ApprovalWorkflows tables"],
    [80, "10. Utilities & Settings", "Audit Trail / Log Management", "COMPLETED", 100, "High", "Done", "AuditLogs, SecurityLogs"],
    [81, "10. Utilities & Settings", "Tax Configuration (GST/VAT)", "COMPLETED", 100, "High", "Done", "Taxes module"],
    [82, "10. Utilities & Settings", "Multi-UOM Settings", "PARTIAL", 60, "Medium", "Partial-Complete", "Units module exists"],
    [83, "10. Utilities & Settings", "Barcode/QR Code Generation", "PARTIAL", 30, "Low", "Partial-Complete", "Field exists, no UI"],
    [84, "10. Utilities & Settings", "Multi-Currency Settings", "NOT STARTED", 0, "Medium", "Pending", "Not implemented"],
    [85, "10. Utilities & Settings", "Financial Year/Period Settings", "NOT STARTED", 0, "Medium", "Pending", "Not implemented"],
    [86, "11. Integration Modules", "CRM Integration", "COMPLETED", 100, "High", "Done", "Full CRM integration"],
    [87, "11. Integration Modules", "Finance/Accounting Integration", "PARTIAL", 20, "High", "Partial-Complete", "Basic expenses only"],
    [88, "11. Integration Modules", "E-commerce/Marketplace Sync", "NOT STARTED", 0, "Low", "Pending", "Not implemented"],
    [89, "11. Integration Modules", "Logistics/Shipping Integration", "NOT STARTED", 0, "Low", "Pending", "Not implemented"],
    [90, "11. Integration Modules", "Barcode Scanner/RFID Integration", "NOT STARTED", 0, "Low", "Pending", "Not implemented"],
    [91, "12. CRM - Core Entities", "Leads Management", "COMPLETED", 100, "High", "Done", "Full leads module"],
    [92, "12. CRM - Core Entities", "Opportunities Management", "COMPLETED", 100, "High", "Done", "Full opportunities module"],
    [93, "12. CRM - Core Entities", "Accounts Management", "COMPLETED", 100, "High", "Done", "Full accounts module"],
    [94, "12. CRM - Core Entities", "Contacts Management", "COMPLETED", 100, "High", "Done", "Full contacts module"],
    [95, "12. CRM - Core Entities", "Activities Management", "COMPLETED", 100, "High", "Done", "Full activities module"],
    [96, "12. CRM - Core Entities", "Cases/Support Tickets", "COMPLETED", 100, "High", "Done", "Full cases module"],
    [97, "12. CRM - Core Entities", "Quotes Management", "COMPLETED", 100, "High", "Done", "Full quotes module"],
    [98, "12. CRM - Core Entities", "Invoices Management", "COMPLETED", 100, "High", "Done", "Full invoices module"],
    [99, "12. CRM - Core Entities", "Payments Management", "COMPLETED", 100, "High", "Done", "Full payments module"],
    [100, "12. CRM - Core Entities", "Retentions/Follow-ups", "COMPLETED", 100, "Medium", "Done", "Full retentions module"],
    [101, "12. CRM - Core Entities", "Presales Management", "PARTIAL", 70, "Medium", "Partial-Complete", "Model exists, UI partial"],
    [102, "13. CRM - Master Data", "Lead Sources", "COMPLETED", 100, "High", "Done", "LeadSources master"],
    [103, "13. CRM - Master Data", "Industries", "COMPLETED", 100, "Medium", "Done", "Industries master"],
    [104, "13. CRM - Master Data", "Sales Stages", "COMPLETED", 100, "High", "Done", "SalesStages master"],
    [105, "13. CRM - Master Data", "Follow-up Types", "COMPLETED", 100, "Medium", "Done", "FollowupTypes master"],
    [106, "13. CRM - Master Data", "Task Types", "COMPLETED", 100, "Medium", "Done", "TaskTypes master"],
    [107, "13. CRM - Master Data", "Payment Terms Master", "NOT STARTED", 0, "Medium", "Pending", "Not implemented"],
    [108, "13. CRM - Master Data", "Case Types/Categories", "NOT STARTED", 0, "Low", "Pending", "Not implemented"],
    [109, "14. CRM - Workflows", "Lead-to-Opportunity Conversion", "COMPLETED", 100, "High", "Done", "Implemented"],
    [110, "14. CRM - Workflows", "Opportunity-to-Quote Conversion", "COMPLETED", 100, "High", "Done", "Implemented"],
    [111, "14. CRM - Workflows", "Quote-to-Invoice Conversion", "COMPLETED", 100, "High", "Done", "Implemented"],
    [112, "14. CRM - Workflows", "Invoice-to-Payment Tracking", "COMPLETED", 100, "High", "Done", "Implemented"],
    [113, "14. CRM - Workflows", "Visual Workflow Builder", "NOT STARTED", 0, "Medium", "Pending", "Not implemented"],
    [114, "14. CRM - Workflows", "Automated Email Notifications", "NOT STARTED", 0, "Medium", "Pending", "Not implemented"],
    [115, "14. CRM - Workflows", "SLA Management", "NOT STARTED", 0, "Low", "Pending", "Not implemented"],
    [116, "15. CRM - Reports & Analytics", "Lead Reports", "COMPLETED", 100, "High", "Done", "Basic reports"],
    [117, "15. CRM - Reports & Analytics", "Opportunity Reports", "COMPLETED", 100, "High", "Done", "Basic reports"],
    [118, "15. CRM - Reports & Analytics", "Sales Pipeline Dashboard", "COMPLETED", 100, "High", "Done", "Dashboard exists"],
    [119, "15. CRM - Reports & Analytics", "Conversion Rate Analytics", "NOT STARTED", 0, "Medium", "Pending", "Not implemented"],
    [120, "15. CRM - Reports & Analytics", "Sales Forecasting", "NOT STARTED", 0, "Medium", "Pending", "Not implemented"],
    [121, "15. CRM - Reports & Analytics", "Lead Source Effectiveness", "NOT STARTED", 0, "Low", "Pending", "Not implemented"],
    [122, "15. CRM - Reports & Analytics", "Custom Report Builder", "NOT STARTED", 0, "Low", "Pending", "Not implemented"],
    [123, "16. CRM - Integration", "Email Integration", "NOT STARTED", 0, "High", "Pending", "Not implemented"],
    [124, "16. CRM - Integration", "Calendar Integration", "NOT STARTED", 0, "Medium", "Pending", "Not implemented"],
    [125, "16. CRM - Integration", "Document Management", "NOT STARTED", 0, "Medium", "Pending", "Not implemented"],
    [126, "16. CRM - Integration", "Customer Portal", "NOT STARTED", 0, "Low", "Pending", "Not implemented"],
    [127, "16. CRM - Integration", "Social Media Integration", "NOT STARTED", 0, "Low", "Pending", "Not implemented"],
    [128, "17. User Management", "User Registration", "COMPLETED", 100, "High", "Done", "RegisterUserPage"],
    [129, "17. User Management", "User Profile Management", "COMPLETED", 100, "High", "Done", "Profile pages"],
    [130, "17. User Management", "User Listing & Search", "COMPLETED", 100, "High", "Done", "UsersPage"],
    [131, "17. User Management", "User Activation/Deactivation", "COMPLETED", 100, "High", "Done", "IsActive flag"],
    [132, "17. User Management", "Profile Image Upload", "COMPLETED", 100, "Medium", "Done", "Implemented"],
    [133, "17. User Management", "Bulk User Import", "PARTIAL", 40, "Medium", "Partial-Complete", "Basic import exists"],
    [134, "17. User Management", "User Impersonation", "NOT STARTED", 0, "Medium", "Pending", "Not implemented"],
    [135, "17. User Management", "User Activity Tracking", "NOT STARTED", 0, "Low", "Pending", "Not implemented"],
    [136, "18. Role Management", "Role Creation & Management", "COMPLETED", 100, "High", "Done", "Roles module"],
    [137, "18. Role Management", "Role Assignment to Users", "COMPLETED", 100, "High", "Done", "RoleId in Users"],
    [138, "18. Role Management", "Predefined Roles", "COMPLETED", 100, "High", "Done", "Owner, Manager, etc."],
    [139, "18. Role Management", "Role Cloning/Duplication", "NOT STARTED", 0, "Low", "Pending", "Not implemented"],
    [140, "18. Role Management", "Role Hierarchy", "NOT STARTED", 0, "Low", "Pending", "Not implemented"],
    [141, "18. Role Management", "Role Templates", "NOT STARTED", 0, "Low", "Pending", "Not implemented"],
    [142, "19. Permission Management", "Permission Structure", "COMPLETED", 100, "High", "Done", "Permissions table"],
    [143, "19. Permission Management", "Role-Level Permissions", "COMPLETED", 100, "High", "Done", "RolePermissions table"],
    [144, "19. Permission Management", "User-Level Permissions", "COMPLETED", 100, "High", "Done", "UserPermissions table"],
    [145, "19. Permission Management", "Field-Level Permissions", "NOT STARTED", 0, "High", "Pending", "Not implemented"],
    [146, "19. Permission Management", "Record-Level Permissions", "NOT STARTED", 0, "High", "Pending", "Not implemented"],
    [147, "19. Permission Management", "Time-Based Permissions", "NOT STARTED", 0, "Low", "Pending", "Not implemented"],
    [148, "19. Permission Management", "Permission Delegation", "NOT STARTED", 0, "Low", "Pending", "Not implemented"],
    [149, "20. Authentication & Security", "Email/Password Authentication", "COMPLETED", 100, "High", "Done", "JWT-based"],
    [150, "20. Authentication & Security", "JWT Token Authentication", "COMPLETED", 100, "High", "Done", "Implemented"],
    [151, "20. Authentication & Security", "Refresh Token Mechanism", "COMPLETED", 100, "High", "Done", "Implemented"],
    [152, "20. Authentication & Security", "Email Verification", "COMPLETED", 100, "High", "Done", "With tokens"],
    [153, "20. Authentication & Security", "Password Reset", "COMPLETED", 100, "High", "Done", "With expiration"],
    [154, "20. Authentication & Security", "Account Locking", "COMPLETED", 100, "High", "Done", "Brute-force protection"],
    [155, "20. Authentication & Security", "Two-Factor Authentication (2FA)", "NOT STARTED", 0, "High", "Pending", "Not implemented"],
    [156, "20. Authentication & Security", "OAuth Integration", "NOT STARTED", 0, "Medium", "Pending", "Not implemented"],
    [157, "20. Authentication & Security", "SSO/SAML Integration", "NOT STARTED", 0, "Medium", "Pending", "Not implemented"],
    [158, "20. Authentication & Security", "LDAP/AD Integration", "NOT STARTED", 0, "Medium", "Pending", "Not implemented"],
    [159, "20. Authentication & Security", "Password Strength Enforcement", "NOT STARTED", 0, "Medium", "Pending", "Not implemented"],
    [160, "20. Authentication & Security", "Password History", "NOT STARTED", 0, "Low", "Pending", "Not implemented"],
    [161, "21. Audit & Compliance", "Audit Trail", "COMPLETED", 100, "High", "Done", "AuditLogs table"],
    [162, "21. Audit & Compliance", "Security Logs", "COMPLETED", 100, "High", "Done", "SecurityLogs table"],
    [163, "21. Audit & Compliance", "Login/Logout Tracking", "COMPLETED", 100, "High", "Done", "Implemented"],
    [164, "21. Audit & Compliance", "Entity-Level Audit Fields", "COMPLETED", 100, "High", "Done", "CreatedBy, UpdatedBy"],
    [165, "21. Audit & Compliance", "Detailed Change Tracking", "NOT STARTED", 0, "Medium", "Pending", "Before/after values"],
    [166, "21. Audit & Compliance", "Audit Log Search & Filter", "NOT STARTED", 0, "Medium", "Pending", "Not implemented"],
    [167, "21. Audit & Compliance", "GDPR Compliance Tools", "NOT STARTED", 0, "Medium", "Pending", "Not implemented"],
    [168, "21. Audit & Compliance", "Data Retention Policies", "NOT STARTED", 0, "Low", "Pending", "Not implemented"],
    [169, "22. Notifications", "In-App Notifications", "COMPLETED", 100, "High", "Done", "Notifications table"],
    [170, "22. Notifications", "Notification Preferences", "COMPLETED", 100, "Medium", "Done", "Preferences exist"],
    [171, "22. Notifications", "Email Notifications", "NOT STARTED", 0, "Medium", "Pending", "Not implemented"],
    [172, "22. Notifications", "SMS Notifications", "NOT STARTED", 0, "Low", "Pending", "Not implemented"],
    [173, "22. Notifications", "Push Notifications", "NOT STARTED", 0, "Low", "Pending", "Not implemented"],
    [174, "22. Notifications", "Real-Time Notifications", "NOT STARTED", 0, "Medium", "Pending", "WebSocket not implemented"],
]

# Populate data with color coding
for row_idx, feature in enumerate(features, start=5):
    for col_idx, value in enumerate(feature, 1):
        cell = ws1.cell(row=row_idx, column=col_idx, value=value)
        cell.alignment = Alignment(vertical='center', wrap_text=True, shrink_to_fit=True)
        cell.font = Font(size=8)
        if feature[6] == "Done":
            cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        elif feature[6] == "Partial-Complete":
            cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        else:
            cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

# ========== SHEET 2: PARTIAL WORK - NEEDS COMPLETION ==========
ws2 = wb.create_sheet("Partial - To Complete")

ws2['A1'] = "PARTIALLY IMPLEMENTED FEATURES - REQUIRE COMPLETION"
ws2['A1'].font = Font(size=14, bold=True, color="FFFFFF")
ws2['A1'].fill = PatternFill(start_color="FF6600", end_color="FF6600", fill_type="solid")
ws2['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws2.merge_cells('A1:F1')
ws2.row_dimensions[1].height = 25

headers2 = ['#', 'Module', 'Feature', 'Completion %', 'Current Status', 'What Needs To Be Done']
for col, header in enumerate(headers2, 1):
    cell = ws2.cell(row=3, column=col, value=header)
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="548235", end_color="548235", fill_type="solid")
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

partial_work = [
    [1, "1. Master Data Management", "Bin/Rack Master", 50, "Basic locations exist", "Dedicated bin/rack management module, assign bins to products"],
    [2, "1. Master Data Management", "Item Barcode Setup", 40, "Barcode field exists", "Barcode generation logic, barcode printing UI, barcode scanning support"],
    [3, "2. Purchase Management", "Purchase Order Approval", 60, "Generic approval system", "PO-specific approval workflow, multi-level approval, approval notifications"],
    [4, "2. Purchase Management", "Purchase Order Tracking", 50, "Basic status tracking", "Detailed tracking, shipment tracking, ETA, timeline view"],
    [5, "3. Sales & Order Management", "Sales Order Approval", 60, "Generic approval system", "SO-specific approval workflow, credit limit check, multi-level approval"],
    [6, "3. Sales & Order Management", "Backorder Management", 40, "Basic handling only", "Backorder creation, allocation, partial shipment, backorder reports"],
    [7, "4. Stock/Inventory Operations", "Stock Reservation/Allocation", 30, "Basic allocation only", "Dedicated reservation system, allocate against SO/PO, release logic"],
    [8, "5. Warehouse Management (WMS)", "Bin/Location Mapping", 50, "Basic warehouse locations", "Detailed bin mapping, zone mapping, putaway rules, picking strategies"],
    [9, "5. Warehouse Management (WMS)", "Packing & Shipment", 40, "Basic delivery challan", "Packing list generation, carrier integration, shipment tracking, label printing"],
    [10, "7. Quality Control", "Quality Inspection Setup", 30, "Model exists, not integrated", "Integration with GRN, inspection checklist, pass/fail logic, quarantine"],
    [11, "8. Production/Manufacturing", "Material Requisition for Production", 50, "Part of production module", "Separate requisition workflow, approval, against WO, tracking"],
    [12, "8. Production/Manufacturing", "Raw Material Issue", 50, "Part of production module", "Dedicated issue workflow, batch picking, cost calculation"],
    [13, "8. Production/Manufacturing", "Finished Goods Receipt", 50, "Part of production module", "QC integration, receipt against WO, automatic inventory update"],
    [14, "9. Reports & Analytics", "Stock Summary Report", 50, "Basic reports exist", "Enhanced filters, export, drill-down, warehouse-wise"],
    [15, "9. Reports & Analytics", "Item-wise Sales & Purchase Analysis", 60, "Basic reporting exists", "Advanced filters, trends, comparison, export"],
    [16, "10. Utilities & Settings", "Multi-UOM Settings", 60, "Units module exists", "UOM conversion rules, auto-conversion on transactions"],
    [17, "10. Utilities & Settings", "Barcode/QR Code Generation", 30, "Field exists, no UI", "Generation service, print templates, scanning integration"],
    [18, "11. Integration Modules", "Finance/Accounting Integration", 20, "Basic expenses only", "Full accounting integration, journal entries, chart of accounts"],
    [19, "12. CRM - Core Entities", "Presales Management", 70, "Model exists, UI partial", "Complete UI, presales workflow, demo scheduling"],
    [20, "17. User Management", "Bulk User Import", 40, "Basic import exists", "Validation, templates, error reporting, role assignment"],
]

for row_idx, item in enumerate(partial_work, start=4):
    for col_idx, value in enumerate(item, 1):
        cell = ws2.cell(row=row_idx, column=col_idx, value=value)
        cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        cell.font = Font(color="9C5700", size=9)
        cell.alignment = Alignment(vertical='center', wrap_text=True)

for col, w in enumerate([6, 30, 35, 14, 30, 50], 1):
    ws2.column_dimensions[openpyxl.utils.cell.get_column_letter(col)].width = w

# ========== SHEET 3: PENDING WORK - NOT STARTED ==========
ws3 = wb.create_sheet("Pending - Not Started")

ws3['A1'] = "NOT STARTED FEATURES - PENDING IMPLEMENTATION"
ws3['A1'].font = Font(size=14, bold=True, color="FFFFFF")
ws3['A1'].fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
ws3['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws3.merge_cells('A1:F1')
ws3.row_dimensions[1].height = 25

headers3 = ['#', 'Module', 'Feature', 'Phase', 'Priority', 'Est. Effort']
for col, header in enumerate(headers3, 1):
    cell = ws3.cell(row=3, column=col, value=header)
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="9C0006", end_color="9C0006", fill_type="solid")
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

pending_work = [
    [1, "6. Stock Valuation & Costing", "Costing Method Setup (FIFO/LIFO)", "1-Critical Core", "HIGH", "3-4 weeks"],
    [2, "6. Stock Valuation & Costing", "Stock Valuation Report", "1-Critical Core", "HIGH", "2-3 weeks"],
    [3, "6. Stock Valuation & Costing", "Landed Cost Allocation", "1-Critical Core", "MEDIUM", "3-4 weeks"],
    [4, "6. Stock Valuation & Costing", "Cost Adjustment Entries", "1-Critical Core", "MEDIUM", "2-3 weeks"],
    [5, "4. Stock/Inventory Operations", "Reorder Level & Auto Replenishment", "1-Critical Core", "HIGH", "3-4 weeks"],
    [6, "4. Stock/Inventory Operations", "Min-Max Stock Settings", "1-Critical Core", "HIGH", "1-2 weeks"],
    [7, "10. Utilities & Settings", "Financial Year/Period Settings", "1-Critical Core", "MEDIUM", "2-3 weeks"],
    [8, "20. Authentication & Security", "Two-Factor Authentication (2FA)", "1-Critical Core", "HIGH", "3-4 weeks"],
    [9, "16. CRM - Integration", "Email Integration", "1-Critical Core", "HIGH", "4-5 weeks"],
    [10, "21. Audit & Compliance", "Detailed Change Tracking", "1-Critical Core", "MEDIUM", "3-4 weeks"],
    [11, "16. CRM - Integration", "Document Management", "1-Critical Core", "MEDIUM", "4-5 weeks"],
    [12, "7. Quality Control", "Incoming Inspection (QC on GRN)", "2-Process Enhancement", "MEDIUM", "3-4 weeks"],
    [13, "7. Quality Control", "Rejected/Quarantine Stock", "2-Process Enhancement", "MEDIUM", "2-3 weeks"],
    [14, "7. Quality Control", "QC Certificate Management", "2-Process Enhancement", "LOW", "2-3 weeks"],
    [15, "2. Purchase Management", "Request for Quotation (RFQ)", "2-Process Enhancement", "MEDIUM", "3-4 weeks"],
    [16, "2. Purchase Management", "Vendor Bill/Invoice Matching", "2-Process Enhancement", "MEDIUM", "4-5 weeks"],
    [17, "9. Reports & Analytics", "ABC Analysis", "2-Process Enhancement", "MEDIUM", "2-3 weeks"],
    [18, "9. Reports & Analytics", "Slow-Moving/Non-Moving Stock Report", "2-Process Enhancement", "MEDIUM", "2-3 weeks"],
    [19, "9. Reports & Analytics", "Reorder Report", "2-Process Enhancement", "HIGH", "1-2 weeks"],
    [20, "9. Reports & Analytics", "Vendor Performance Report", "2-Process Enhancement", "MEDIUM", "2-3 weeks"],
    [21, "9. Reports & Analytics", "Stock Aging Report", "2-Process Enhancement", "MEDIUM", "2-3 weeks"],
    [22, "9. Reports & Analytics", "GRN vs PO Reconciliation", "2-Process Enhancement", "MEDIUM", "2-3 weeks"],
    [23, "15. CRM - Reports & Analytics", "Conversion Rate Analytics", "2-Process Enhancement", "MEDIUM", "2-3 weeks"],
    [24, "15. CRM - Reports & Analytics", "Sales Forecasting", "2-Process Enhancement", "MEDIUM", "3-4 weeks"],
    [25, "15. CRM - Reports & Analytics", "Lead Source Effectiveness", "2-Process Enhancement", "LOW", "2 weeks"],
    [26, "15. CRM - Reports & Analytics", "Custom Report Builder", "2-Process Enhancement", "LOW", "4-5 weeks"],
    [27, "13. CRM - Master Data", "Payment Terms Master", "2-Process Enhancement", "MEDIUM", "1 week"],
    [28, "13. CRM - Master Data", "Case Types/Categories", "2-Process Enhancement", "LOW", "1 week"],
    [29, "1. Master Data Management", "HSN/SAC Code Master", "2-Process Enhancement", "MEDIUM", "2-3 weeks"],
    [30, "1. Master Data Management", "Price List Master", "2-Process Enhancement", "MEDIUM", "3-4 weeks"],
    [31, "14. CRM - Workflows", "Visual Workflow Builder", "2-Process Enhancement", "MEDIUM", "4-5 weeks"],
    [32, "14. CRM - Workflows", "Automated Email Notifications", "2-Process Enhancement", "MEDIUM", "3-4 weeks"],
    [33, "14. CRM - Workflows", "SLA Management", "2-Process Enhancement", "LOW", "2-3 weeks"],
    [34, "4. Stock/Inventory Operations", "Expiry Date Management", "2-Process Enhancement", "MEDIUM", "2-3 weeks"],
    [35, "5. Warehouse Management (WMS)", "Putaway Management", "3-Advanced Features", "LOW", "3-4 weeks"],
    [36, "5. Warehouse Management (WMS)", "Picking List Generation", "3-Advanced Features", "LOW", "3-4 weeks"],
    [37, "5. Warehouse Management (WMS)", "Cycle Count", "3-Advanced Features", "LOW", "3-4 weeks"],
    [38, "5. Warehouse Management (WMS)", "Physical Stock Verification", "3-Advanced Features", "LOW", "3-4 weeks"],
    [39, "8. Production/Manufacturing", "By-product/Scrap Entry", "3-Advanced Features", "LOW", "2-3 weeks"],
    [40, "19. Permission Management", "Field-Level Permissions", "3-Advanced Features", "HIGH", "3-4 weeks"],
    [41, "19. Permission Management", "Record-Level Permissions", "3-Advanced Features", "HIGH", "3-4 weeks"],
    [42, "19. Permission Management", "Time-Based Permissions", "3-Advanced Features", "LOW", "2-3 weeks"],
    [43, "19. Permission Management", "Permission Delegation", "3-Advanced Features", "LOW", "2 weeks"],
    [44, "17. User Management", "User Impersonation", "3-Advanced Features", "MEDIUM", "1-2 weeks"],
    [45, "17. User Management", "User Activity Tracking", "3-Advanced Features", "LOW", "2-3 weeks"],
    [46, "18. Role Management", "Role Cloning/Duplication", "3-Advanced Features", "LOW", "1 week"],
    [47, "18. Role Management", "Role Hierarchy", "3-Advanced Features", "LOW", "2-3 weeks"],
    [48, "18. Role Management", "Role Templates", "3-Advanced Features", "LOW", "2 weeks"],
    [49, "20. Authentication & Security", "OAuth Integration", "3-Advanced Features", "MEDIUM", "2-3 weeks"],
    [50, "20. Authentication & Security", "SSO/SAML Integration", "3-Advanced Features", "MEDIUM", "3-4 weeks"],
    [51, "20. Authentication & Security", "LDAP/AD Integration", "3-Advanced Features", "MEDIUM", "3-4 weeks"],
    [52, "20. Authentication & Security", "Password Strength Enforcement", "3-Advanced Features", "MEDIUM", "1 week"],
    [53, "20. Authentication & Security", "Password History", "3-Advanced Features", "LOW", "1 week"],
    [54, "21. Audit & Compliance", "Audit Log Search & Filter", "3-Advanced Features", "MEDIUM", "2-3 weeks"],
    [55, "21. Audit & Compliance", "GDPR Compliance Tools", "3-Advanced Features", "MEDIUM", "3-4 weeks"],
    [56, "21. Audit & Compliance", "Data Retention Policies", "3-Advanced Features", "LOW", "2 weeks"],
    [57, "10. Utilities & Settings", "Multi-Currency Settings", "3-Advanced Features", "MEDIUM", "3-4 weeks"],
    [58, "2. Purchase Management", "Blanket/Contract Orders", "4-Integration", "LOW", "3-4 weeks"],
    [59, "5. Warehouse Management (WMS)", "Cross-Docking", "4-Integration", "LOW", "3-4 weeks"],
    [60, "11. Integration Modules", "E-commerce/Marketplace Sync", "4-Integration", "LOW", "4-5 weeks"],
    [61, "11. Integration Modules", "Logistics/Shipping Integration", "4-Integration", "LOW", "4-5 weeks"],
    [62, "11. Integration Modules", "Barcode Scanner/RFID Integration", "4-Integration", "LOW", "2-3 weeks"],
    [63, "19. Marketing Automation", "Campaign Management", "4-Integration", "LOW", "4-5 weeks"],
    [64, "20. Workflow Automation", "Visual Workflow Builder", "4-Integration", "LOW", "5-6 weeks"],
    [65, "21. Field-Level Security", "Permission Matrix UI", "4-Integration", "LOW", "2-3 weeks"],
    [66, "22. User Experience", "Dashboard Customization", "4-Integration", "LOW", "3-4 weeks"],
    [67, "22. User Experience", "Saved Views/Filters", "4-Integration", "LOW", "2-3 weeks"],
    [68, "22. User Experience", "Keyboard Shortcuts", "4-Integration", "LOW", "1 week"],
    [69, "22. User Experience", "User Preferences", "4-Integration", "LOW", "2-3 weeks"],
    [70, "22. Notifications", "Email Notifications", "4-Integration", "MEDIUM", "2-3 weeks"],
    [71, "22. Notifications", "SMS Notifications", "4-Integration", "LOW", "2-3 weeks"],
    [72, "22. Notifications", "Push Notifications", "4-Integration", "LOW", "2-3 weeks"],
    [73, "22. Notifications", "Real-Time Notifications", "4-Integration", "MEDIUM", "2-3 weeks"],
    [74, "23. Multi-Currency Support", "Exchange Rate Management", "4-Integration", "LOW", "2-3 weeks"],
    [75, "23. Multi-Currency Support", "Currency Conversion Logic", "4-Integration", "LOW", "3-4 weeks"],
    [76, "23. Multi-Currency Support", "Gain/Loss Calculation", "4-Integration", "LOW", "2-3 weeks"],
    [77, "24. Chart of Accounts", "Account Hierarchy", "4-Integration", "LOW", "2-3 weeks"],
    [78, "24. Chart of Accounts", "Journal Entry System", "4-Integration", "LOW", "4-5 weeks"],
    [79, "24. Chart of Accounts", "Trial Balance", "4-Integration", "LOW", "2-3 weeks"],
    [80, "25. Calendar Integration", "Google Calendar Sync", "4-Integration", "LOW", "2-3 weeks"],
    [81, "25. Calendar Integration", "Outlook Calendar Sync", "4-Integration", "LOW", "2-3 weeks"],
    [82, "26. SMS & Push Notifications", "SMS Gateway Setup", "4-Integration", "LOW", "2 weeks"],
    [83, "26. SMS & Push Notifications", "Notification Templates", "4-Integration", "LOW", "1 week"],
    [84, "27. Social Media Integration", "LinkedIn API", "4-Integration", "LOW", "2-3 weeks"],
    [85, "27. Social Media Integration", "Twitter/X API", "4-Integration", "LOW", "2-3 weeks"],
    [86, "27. Social Media Integration", "Facebook API", "4-Integration", "LOW", "2-3 weeks"],
    [87, "28. Customer Portal", "Customer Authentication", "4-Integration", "LOW", "2-3 weeks"],
    [88, "28. Customer Portal", "Self-Service Features", "4-Integration", "LOW", "3-4 weeks"],
    [89, "28. Customer Portal", "Order Tracking", "4-Integration", "LOW", "2-3 weeks"],
    [90, "29. E-commerce Integration", "Shopify/WooCommerce API", "4-Integration", "LOW", "3-4 weeks"],
    [91, "30. Shipping Integration", "Shipping APIs (FedEx/UPS/DHL)", "4-Integration", "LOW", "3-4 weeks"],
]

for row_idx, item in enumerate(pending_work, start=4):
    for col_idx, value in enumerate(item, 1):
        cell = ws3.cell(row=row_idx, column=col_idx, value=value)
        cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        cell.font = Font(color="9C0006", size=9)
        cell.alignment = Alignment(vertical='center', wrap_text=True)

for col, w in enumerate([6, 35, 40, 22, 12, 20], 1):
    ws3.column_dimensions[openpyxl.utils.cell.get_column_letter(col)].width = w

# ========== SHEET 4: IMPLEMENTATION PLAN ==========
ws4 = wb.create_sheet("Implementation Plan")

ws4['A1'] = "RECOMMENDED IMPLEMENTATION PLAN"
ws4['A1'].font = Font(size=14, bold=True, color="FFFFFF")
ws4['A1'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
ws4['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws4.merge_cells('A1:F1')
ws4.row_dimensions[1].height = 25

headers4 = ['Phase', 'Module', 'Feature', 'Action Required', 'Est. Time', 'Dev Count']
for col, header in enumerate(headers4, 1):
    cell = ws4.cell(row=3, column=col, value=header)
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

plan = [
    ["PHASE 1 (Months 1-2): Complete Partial Work", "Multiple", "All 20 Partial Features", "Complete implementation of all partial features", "8-10 weeks", "3 developers"],
    ["PHASE 1 (Months 3-4): Critical Core Features", "6. Stock Valuation", "Full Costing Module", "Implement tables, models, calculations, reports", "4 weeks", "2 developers"],
    ["PHASE 1 (Months 3-4): Critical Core Features", "4. Stock/Inventory", "Reorder Management", "Reorder tables, auto-replenishment, alerts", "3 weeks", "1 developer"],
    ["PHASE 1 (Months 3-4): Critical Core Features", "20. Authentication", "2FA Implementation", "TOTP, backup codes, login flow update", "3 weeks", "1 developer"],
    ["PHASE 1 (Months 3-4): Critical Core Features", "10. Utilities", "Financial Year Settings", "Financial years, periods, closing logic", "2 weeks", "1 developer"],
    ["PHASE 1 (Months 3-4): Critical Core Features", "16. CRM Integration", "Email Integration", "Email service, templates, tracking, sync", "4 weeks", "1 developer"],
    ["PHASE 2 (Months 5-6): Process Enhancement", "7. Quality Control", "Full QC Module", "QC workflow, inspection, certificates", "5 weeks", "1 developer"],
    ["PHASE 2 (Months 5-6): Process Enhancement", "2. Purchase Management", "RFQ & 3-Way Match", "RFQ workflow, vendor matching, variance detection", "6 weeks", "2 developers"],
    ["PHASE 2 (Months 6-7): Process Enhancement", "9. Reports & Analytics", "Advanced Reports", "Aging, ABC, slow-moving, vendor performance", "4 weeks", "1 developer"],
    ["PHASE 2 (Months 7-8): Process Enhancement", "13. Master Data", "Price List & HSN/SAC", "Price lists, customer/supplier pricing, HSN/SAC master", "4 weeks", "1 developer"],
    ["PHASE 2 (Months 7-8): Process Enhancement", "14. CRM Workflows", "Advanced Workflows", "Visual workflow builder, automation rules", "5 weeks", "1 developer"],
    ["PHASE 3 (Months 9-10): Advanced Features", "5. WMS", "Putaway & Picking", "Putaway rules, picking lists, wave picking", "6 weeks", "1 developer"],
    ["PHASE 3 (Months 10-11): Advanced Features", "19. Permissions", "Advanced Security", "Field-level, record-level, dynamic permissions", "6 weeks", "1 developer"],
    ["PHASE 3 (Months 11-12): Advanced Features", "17. User/Role Mgmt", "User Experience", "Role hierarchy, impersonation, activity tracking", "4 weeks", "1 developer"],
    ["PHASE 3 (Months 11-12): Advanced Features", "20. Authentication", "Enterprise Auth", "OAuth, SSO/SAML, LDAP, password policies", "5 weeks", "1 developer"],
    ["PHASE 4 (Months 13-14): Integration", "11. Integration", "E-commerce & Shipping", "Shopify, WooCommerce, FedEx/UPS/DHL", "6 weeks", "1 developer"],
    ["PHASE 4 (Months 14-15): Integration", "28. Customer Portal", "Self-Service Portal", "Customer auth, order tracking, invoices", "6 weeks", "1 developer"],
    ["PHASE 4 (Months 15-16): Integration", "24. Accounting", "Chart of Accounts", "COA, journal entries, trial balance", "6 weeks", "2 developers"],
]

for row_idx, item in enumerate(plan, start=4):
    for col_idx, value in enumerate(item, 1):
        cell = ws4.cell(row=row_idx, column=col_idx, value=value)
        cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        cell.font = Font(size=9)
        cell.alignment = Alignment(vertical='center', wrap_text=True)

for col, w in enumerate([35, 25, 30, 45, 15, 16], 1):
    ws4.column_dimensions[openpyxl.utils.cell.get_column_letter(col)].width = w

# ========== SHEET 5: SUMMARY ==========
ws5 = wb.create_sheet("Summary")

ws5['A1'] = "PROJECT SUMMARY & STATISTICS"
ws5['A1'].font = Font(size=14, bold=True, color="FFFFFF")
ws5['A1'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
ws5['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws5.merge_cells('A1:C1')
ws5.row_dimensions[1].height = 25

summary_data = [
    ["", "", ""],
    ["OVERALL PROJECT STATUS", "", ""],
    ["Total Features", 174, ""],
    ["Completed", 82, "47.1%"],
    ["Partial (Need Completion)", 20, "11.5%"],
    ["Not Started", 72, "41.4%"],
    ["Overall Completion (Weighted)", "52.9%", ""],
    ["", "", ""],
    ["BY CATEGORY", "", ""],
    ["Done (Ready for Use)", 82, ""],
    ["Partial -> Complete", 20, "Need finishing"],
    ["Pending (Not Started)", 72, ""],
    ["", "", ""],
    ["BY PRIORITY", "", ""],
    ["HIGH Priority Pending", "~25 features", "Do first"],
    ["MEDIUM Priority Pending", "~32 features", "Do second"],
    ["LOW Priority Pending", "~15 features", "Do last"],
    ["", "", ""],
    ["BY PHASE", "", ""],
    ["Phase 1 (Months 1-4)", "Critical Core", "HIGH priority"],
    ["Phase 2 (Months 5-8)", "Process Enhancement", "MEDIUM priority"],
    ["Phase 3 (Months 9-12)", "Advanced Features", "LOW priority"],
    ["Phase 4 (Months 13-16)", "Integration & Extensions", "LOW priority"],
    ["", "", ""],
    ["TEAM REQUIREMENTS", "", ""],
    ["Recommended Team Size", "3-5 developers + 1 QA + 1 UI/UX", ""],
    ["Estimated Timeline", "12-16 months", ""],
    ["", "", ""],
    ["IMMEDIATE ACTIONS (First 2 Weeks)", "", ""],
    ["1. Complete Partial Features", "20 features", "Highest ROI"],
    ["2. Start Stock Valuation", "High priority", "Business critical"],
    ["3. Start Reorder Management", "High priority", "Business critical"],
    ["4. Implement 2FA", "Security requirement", ""],
]

for row_idx, item in enumerate(summary_data, start=3):
    ws5.cell(row=row_idx, column=1, value=item[0])
    ws5.cell(row=row_idx, column=2, value=item[1])
    ws5.cell(row=row_idx, column=3, value=item[2])
    if item[0] and item[0] in ["OVERALL PROJECT STATUS", "BY CATEGORY", "BY PRIORITY", "BY PHASE", "TEAM REQUIREMENTS", "IMMEDIATE ACTIONS (First 2 Weeks)"]:
        fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        font = Font(bold=True, color="FFFFFF", size=11)
    elif item[0] == "":
        fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        font = Font(size=10)
    else:
        fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        font = Font(size=10)
    for col in range(1, 4):
        ws5.cell(row=row_idx, column=col).fill = fill
        ws5.cell(row=row_idx, column=col).font = font
        ws5.cell(row=row_idx, column=col).alignment = Alignment(vertical='center', wrap_text=True)

for col, w in enumerate([35, 40, 30], 1):
    ws5.column_dimensions[openpyxl.utils.cell.get_column_letter(col)].width = w

# Save workbook
wb.save("ERP_CRM_Done_and_Pending_Work.xlsx")
print("Excel file created: ERP_CRM_Done_and_Pending_Work.xlsx")
print("Sheets: Feature Status, Partial - To Complete, Pending - Not Started, Implementation Plan, Summary")
