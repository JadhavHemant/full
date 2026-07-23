import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Create workbook
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Project Completion Status"

# Define styles
header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
header_font = Font(color="FFFFFF", bold=True, size=11)
complete_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
partial_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
not_started_fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Title
ws.merge_cells('A1:G1')
ws['A1'] = "ERP/CRM Project Completion Status Report"
ws['A1'].font = Font(size=16, bold=True)
ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws.row_dimensions[1].height = 30

# Subtitle
ws.merge_cells('A2:G2')
ws['A2'] = "Generated: 2026-07-22 | Total Features Analyzed: 150+"
ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
ws['A2'].font = Font(size=10, italic=True)

# Headers
headers = ['Module', 'Feature', 'Status', 'Completion %', 'Priority', 'Phase', 'Notes']
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=4, column=col, value=header)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = thin_border

# Data structure: [Module, Feature, Status, Completion%, Priority, Phase, Notes]
data = [
    # 1. Master Data Management
    ["1. Master Data Management", "Item/Product Master", "Complete", 100, "High", "1", "Full CRUD with all fields"],
    ["1. Master Data Management", "Item Category / Sub-Category", "Complete", 100, "High", "1", "ProductCategories module"],
    ["1. Master Data Management", "Unit of Measurement (UOM)", "Complete", 100, "High", "1", "Units module with CRUD"],
    ["1. Master Data Management", "Warehouse/Location Master", "Complete", 100, "High", "1", "Warehouses module"],
    ["1. Master Data Management", "Bin/Rack Master", "Partial", 50, "Medium", "2", "Basic locations exist"],
    ["1. Master Data Management", "Supplier/Vendor Master", "Complete", 100, "High", "1", "Suppliers module"],
    ["1. Master Data Management", "Customer Master", "Complete", 100, "High", "1", "Customers module"],
    ["1. Master Data Management", "Brand Master", "Complete", 100, "Medium", "1", "Brands module"],
    ["1. Master Data Management", "Item Barcode Setup", "Partial", 40, "Low", "3", "Barcode field exists"],
    ["1. Master Data Management", "HSN/SAC Code Master", "Not Started", 0, "Medium", "2", "Not implemented"],
    ["1. Master Data Management", "Item Attributes (Size, Color)", "Not Started", 0, "Low", "3", "Only Batch/Serial tracked"],
    ["1. Master Data Management", "Item Attributes (Batch, Serial)", "Complete", 100, "High", "1", "BatchSerial module"],
    ["1. Master Data Management", "Price List Master", "Not Started", 0, "Medium", "2", "Not implemented"],
    ["1. Master Data Management", "Tax Configuration", "Complete", 100, "High", "1", "Taxes and ProductTaxMap"],

    # 2. Purchase Management
    ["2. Purchase Management", "Purchase Requisition (PR)", "Complete", 100, "High", "1", "PurchaseRequisitions module"],
    ["2. Purchase Management", "Request for Quotation (RFQ)", "Not Started", 0, "Medium", "2", "Not implemented"],
    ["2. Purchase Management", "Purchase Order (PO)", "Complete", 100, "High", "1", "PurchaseOrders with items"],
    ["2. Purchase Management", "Purchase Order Approval", "Partial", 60, "High", "1", "Generic approval system"],
    ["2. Purchase Management", "Goods Receipt Note (GRN)", "Complete", 100, "High", "1", "GRN module"],
    ["2. Purchase Management", "Purchase Return / Debit Note", "Complete", 100, "High", "1", "PurchaseReturns module"],
    ["2. Purchase Management", "Vendor Bill/Invoice Matching", "Not Started", 0, "High", "2", "No 3-way matching"],
    ["2. Purchase Management", "Purchase Order Tracking", "Partial", 50, "Medium", "2", "Basic status tracking"],
    ["2. Purchase Management", "Blanket/Contract Orders", "Not Started", 0, "Low", "3", "Not implemented"],

    # 3. Sales & Order Management
    ["3. Sales & Order Management", "Sales Quotation", "Complete", 100, "High", "1", "SalesQuotations module"],
    ["3. Sales & Order Management", "Sales Order (SO)", "Complete", 100, "High", "1", "SalesOrders module"],
    ["3. Sales & Order Management", "Sales Order Approval", "Partial", 60, "High", "1", "Generic approval system"],
    ["3. Sales & Order Management", "Delivery Challan / Dispatch Note", "Complete", 100, "High", "1", "DeliveryChallans module"],
    ["3. Sales & Order Management", "Sales Return / Credit Note", "Complete", 100, "High", "1", "SalesReturns module"],
    ["3. Sales & Order Management", "Invoice Generation", "Complete", 100, "High", "1", "Invoices module (CRM)"],
    ["3. Sales & Order Management", "Backorder Management", "Partial", 40, "Medium", "2", "Basic handling only"],

    # 4. Stock/Inventory Operations
    ["4. Stock/Inventory Operations", "Stock Inward Entry", "Complete", 100, "High", "1", "Part of GRN process"],
    ["4. Stock/Inventory Operations", "Stock Outward Entry", "Complete", 100, "High", "1", "Part of DeliveryChallan"],
    ["4. Stock/Inventory Operations", "Stock Transfer (Inter-warehouse)", "Complete", 100, "High", "1", "StockTransfers module"],
    ["4. Stock/Inventory Operations", "Stock Adjustment (Positive/Negative)", "Complete", 100, "High", "1", "StockAdjustments module"],
    ["4. Stock/Inventory Operations", "Stock Reservation/Allocation", "Partial", 30, "Medium", "2", "Basic allocation only"],
    ["4. Stock/Inventory Operations", "Batch/Lot Tracking", "Complete", 100, "High", "1", "BatchSerial module"],
    ["4. Stock/Inventory Operations", "Serial Number Tracking", "Complete", 100, "High", "1", "BatchSerial module"],
    ["4. Stock/Inventory Operations", "Expiry Date Management", "Not Started", 0, "Medium", "2", "Not implemented"],
    ["4. Stock/Inventory Operations", "Stock Aging Report", "Not Started", 0, "Medium", "2", "Not implemented"],
    ["4. Stock/Inventory Operations", "Reorder Level & Auto Replenishment", "Not Started", 0, "High", "1", "Not implemented"],
    ["4. Stock/Inventory Operations", "Min-Max Stock Settings", "Not Started", 0, "High", "1", "Not implemented"],
    ["4. Stock/Inventory Operations", "Stock Movements", "Complete", 100, "High", "1", "StockMovements module"],

    # 5. Warehouse Management (WMS)
    ["5. Warehouse Management (WMS)", "Bin/Location Mapping", "Partial", 50, "Medium", "2", "Basic warehouse locations"],
    ["5. Warehouse Management (WMS)", "Putaway Management", "Not Started", 0, "Medium", "3", "Not implemented"],
    ["5. Warehouse Management (WMS)", "Picking List Generation", "Not Started", 0, "Medium", "3", "Not implemented"],
    ["5. Warehouse Management (WMS)", "Packing & Shipment", "Partial", 40, "Medium", "2", "Basic delivery challan"],
    ["5. Warehouse Management (WMS)", "Cycle Count", "Not Started", 0, "Low", "3", "Not implemented"],
    ["5. Warehouse Management (WMS)", "Physical Stock Verification", "Not Started", 0, "Low", "3", "Not implemented"],
    ["5. Warehouse Management (WMS)", "Warehouse Transfer Note", "Complete", 100, "High", "1", "StockTransfers module"],
    ["5. Warehouse Management (WMS)", "Cross-Docking", "Not Started", 0, "Low", "4", "Not implemented"],

    # 6. Stock Valuation & Costing
    ["6. Stock Valuation & Costing", "Costing Method Setup (FIFO/LIFO/Weighted Avg)", "Not Started", 0, "High", "1", "Not implemented"],
    ["6. Stock Valuation & Costing", "Stock Valuation Report", "Not Started", 0, "High", "1", "Not implemented"],
    ["6. Stock Valuation & Costing", "Landed Cost Allocation", "Not Started", 0, "Medium", "2", "Not implemented"],
    ["6. Stock Valuation & Costing", "Cost Adjustment Entries", "Not Started", 0, "Medium", "2", "Not implemented"],

    # 7. Quality Control
    ["7. Quality Control", "Quality Inspection Setup", "Partial", 30, "Medium", "2", "Model exists, not integrated"],
    ["7. Quality Control", "Incoming Inspection (QC on GRN)", "Not Started", 0, "Medium", "2", "Not integrated with GRN"],
    ["7. Quality Control", "Rejected/Quarantine Stock", "Not Started", 0, "Medium", "2", "Not implemented"],
    ["7. Quality Control", "QC Certificate Management", "Not Started", 0, "Low", "3", "Not implemented"],

    # 8. Production/Manufacturing
    ["8. Production/Manufacturing", "Bill of Materials (BOM)", "Complete", 100, "High", "1", "BOM module"],
    ["8. Production/Manufacturing", "Material Requisition for Production", "Partial", 50, "Medium", "2", "Part of production module"],
    ["8. Production/Manufacturing", "Work Order / Job Card", "Complete", 100, "High", "1", "ProductionOrders module"],
    ["8. Production/Manufacturing", "Raw Material Issue", "Partial", 50, "Medium", "2", "Part of production module"],
    ["8. Production/Manufacturing", "Finished Goods Receipt", "Partial", 50, "Medium", "2", "Part of production module"],
    ["8. Production/Manufacturing", "By-product/Scrap Entry", "Not Started", 0, "Low", "3", "Not implemented"],

    # 9. Reports & Analytics
    ["9. Reports & Analytics", "Stock Summary Report", "Partial", 50, "Medium", "2", "Basic reports exist"],
    ["9. Reports & Analytics", "Stock Ledger / Movement Report", "Complete", 100, "High", "1", "StockMovements module"],
    ["9. Reports & Analytics", "Stock Valuation Report", "Not Started", 0, "High", "1", "Not implemented"],
    ["9. Reports & Analytics", "ABC Analysis", "Not Started", 0, "Medium", "3", "Not implemented"],
    ["9. Reports & Analytics", "Slow-Moving/Non-Moving Stock Report", "Not Started", 0, "Medium", "3", "Not implemented"],
    ["9. Reports & Analytics", "Reorder Report", "Not Started", 0, "High", "1", "Not implemented"],
    ["9. Reports & Analytics", "Vendor Performance Report", "Not Started", 0, "Medium", "3", "Not implemented"],
    ["9. Reports & Analytics", "Item-wise Sales & Purchase Analysis", "Partial", 60, "Medium", "2", "Basic reporting exists"],
    ["9. Reports & Analytics", "Batch/Serial Traceability Report", "Complete", 100, "High", "1", "BatchSerial module"],
    ["9. Reports & Analytics", "Stock Aging Report", "Not Started", 0, "Medium", "2", "Not implemented"],
    ["9. Reports & Analytics", "GRN vs PO Reconciliation", "Not Started", 0, "Medium", "2", "Not implemented"],
    ["9. Reports & Analytics", "Profit/Loss Reports", "Complete", 100, "High", "1", "ProfitLossReports module"],

    # 10. Utilities & Settings
    ["10. Utilities & Settings", "Barcode/QR Code Generation", "Partial", 30, "Low", "3", "Field exists, no UI"],
    ["10. Utilities & Settings", "Import/Export Data (Excel/CSV)", "Complete", 100, "High", "1", "DataImportExport module"],
    ["10. Utilities & Settings", "User Role & Permission Setup", "Complete", 100, "High", "1", "Full RBAC system"],
    ["10. Utilities & Settings", "Approval Workflow Configuration", "Complete", 100, "High", "1", "ApprovalWorkflows tables"],
    ["10. Utilities & Settings", "Audit Trail / Log Management", "Complete", 100, "High", "1", "AuditLogs, SecurityLogs"],
    ["10. Utilities & Settings", "Multi-Currency Settings", "Not Started", 0, "Medium", "2", "Not implemented"],
    ["10. Utilities & Settings", "Multi-UOM Settings", "Partial", 60, "Medium", "2", "Units module exists"],
    ["10. Utilities & Settings", "Tax Configuration (GST/VAT)", "Complete", 100, "High", "1", "Taxes module"],
    ["10. Utilities & Settings", "Financial Year/Period Settings", "Not Started", 0, "Medium", "2", "Not implemented"],

    # 11. Integration Modules
    ["11. Integration Modules", "Finance/Accounting Integration", "Partial", 20, "High", "2", "Basic expenses only"],
    ["11. Integration Modules", "CRM Integration", "Complete", 100, "High", "1", "Full CRM integration"],
    ["11. Integration Modules", "E-commerce/Marketplace Sync", "Not Started", 0, "Low", "4", "Not implemented"],
    ["11. Integration Modules", "Logistics/Shipping Integration", "Not Started", 0, "Low", "4", "Not implemented"],
    ["11. Integration Modules", "Barcode Scanner/RFID Integration", "Not Started", 0, "Low", "4", "Not implemented"],

    # 12. CRM Module - Core Entities
    ["12. CRM - Core Entities", "Leads Management", "Complete", 100, "High", "1", "Full leads module"],
    ["12. CRM - Core Entities", "Opportunities Management", "Complete", 100, "High", "1", "Full opportunities module"],
    ["12. CRM - Core Entities", "Accounts Management", "Complete", 100, "High", "1", "Full accounts module"],
    ["12. CRM - Core Entities", "Contacts Management", "Complete", 100, "High", "1", "Full contacts module"],
    ["12. CRM - Core Entities", "Activities Management", "Complete", 100, "High", "1", "Full activities module"],
    ["12. CRM - Core Entities", "Cases/Support Tickets", "Complete", 100, "High", "1", "Full cases module"],
    ["12. CRM - Core Entities", "Quotes Management", "Complete", 100, "High", "1", "Full quotes module"],
    ["12. CRM - Core Entities", "Invoices Management", "Complete", 100, "High", "1", "Full invoices module"],
    ["12. CRM - Core Entities", "Payments Management", "Complete", 100, "High", "1", "Full payments module"],
    ["12. CRM - Core Entities", "Retentions/Follow-ups", "Complete", 100, "Medium", "1", "Full retentions module"],
    ["12. CRM - Core Entities", "Presales Management", "Partial", 70, "Medium", "2", "Model exists, UI partial"],

    # 13. CRM - Master Data
    ["13. CRM - Master Data", "Lead Sources", "Complete", 100, "High", "1", "LeadSources master"],
    ["13. CRM - Master Data", "Industries", "Complete", 100, "Medium", "1", "Industries master"],
    ["13. CRM - Master Data", "Sales Stages", "Complete", 100, "High", "1", "SalesStages master"],
    ["13. CRM - Master Data", "Follow-up Types", "Complete", 100, "Medium", "1", "FollowupTypes master"],
    ["13. CRM - Master Data", "Task Types", "Complete", 100, "Medium", "1", "TaskTypes master"],
    ["13. CRM - Master Data", "Payment Terms Master", "Not Started", 0, "Medium", "2", "Not implemented"],
    ["13. CRM - Master Data", "Case Types/Categories", "Not Started", 0, "Low", "3", "Not implemented"],

    # 14. CRM - Workflows & Automation
    ["14. CRM - Workflows", "Lead-to-Opportunity Conversion", "Complete", 100, "High", "1", "Implemented"],
    ["14. CRM - Workflows", "Opportunity-to-Quote Conversion", "Complete", 100, "High", "1", "Implemented"],
    ["14. CRM - Workflows", "Quote-to-Invoice Conversion", "Complete", 100, "High", "1", "Implemented"],
    ["14. CRM - Workflows", "Invoice-to-Payment Tracking", "Complete", 100, "High", "1", "Implemented"],
    ["14. CRM - Workflows", "Visual Workflow Builder", "Not Started", 0, "Medium", "3", "Not implemented"],
    ["14. CRM - Workflows", "Automated Email Notifications", "Not Started", 0, "Medium", "2", "Not implemented"],
    ["14. CRM - Workflows", "SLA Management", "Not Started", 0, "Low", "3", "Not implemented"],

    # 15. CRM - Reports & Analytics
    ["15. CRM - Reports & Analytics", "Lead Reports", "Complete", 100, "High", "1", "Basic reports"],
    ["15. CRM - Reports & Analytics", "Opportunity Reports", "Complete", 100, "High", "1", "Basic reports"],
    ["15. CRM - Reports & Analytics", "Sales Pipeline Dashboard", "Complete", 100, "High", "1", "Dashboard exists"],
    ["15. CRM - Reports & Analytics", "Conversion Rate Analytics", "Not Started", 0, "Medium", "2", "Not implemented"],
    ["15. CRM - Reports & Analytics", "Sales Forecasting", "Not Started", 0, "Medium", "2", "Not implemented"],
    ["15. CRM - Reports & Analytics", "Lead Source Effectiveness", "Not Started", 0, "Low", "3", "Not implemented"],
    ["15. CRM - Reports & Analytics", "Custom Report Builder", "Not Started", 0, "Low", "3", "Not implemented"],

    # 16. CRM - Integration
    ["16. CRM - Integration", "Email Integration", "Not Started", 0, "High", "2", "Not implemented"],
    ["16. CRM - Integration", "Calendar Integration", "Not Started", 0, "Medium", "2", "Not implemented"],
    ["16. CRM - Integration", "Document Management", "Not Started", 0, "Medium", "2", "Not implemented"],
    ["16. CRM - Integration", "Customer Portal", "Not Started", 0, "Low", "3", "Not implemented"],
    ["16. CRM - Integration", "Social Media Integration", "Not Started", 0, "Low", "4", "Not implemented"],

    # 17. User Management
    ["17. User Management", "User Registration", "Complete", 100, "High", "1", "RegisterUserPage"],
    ["17. User Management", "User Profile Management", "Complete", 100, "High", "1", "Profile pages"],
    ["17. User Management", "User Listing & Search", "Complete", 100, "High", "1", "UsersPage"],
    ["17. User Management", "User Activation/Deactivation", "Complete", 100, "High", "1", "IsActive flag"],
    ["17. User Management", "Profile Image Upload", "Complete", 100, "Medium", "1", "Implemented"],
    ["17. User Management", "User Impersonation", "Not Started", 0, "Medium", "2", "Not implemented"],
    ["17. User Management", "User Activity Tracking", "Not Started", 0, "Low", "3", "Not implemented"],
    ["17. User Management", "Bulk User Import", "Partial", 40, "Medium", "2", "Basic import exists"],

    # 18. Role Management
    ["18. Role Management", "Role Creation & Management", "Complete", 100, "High", "1", "Roles module"],
    ["18. Role Management", "Role Assignment to Users", "Complete", 100, "High", "1", "RoleId in Users"],
    ["18. Role Management", "Predefined Roles", "Complete", 100, "High", "1", "Owner, Manager, etc."],
    ["18. Role Management", "Role Cloning/Duplication", "Not Started", 0, "Low", "3", "Not implemented"],
    ["18. Role Management", "Role Hierarchy", "Not Started", 0, "Low", "3", "Not implemented"],
    ["18. Role Management", "Role Templates", "Not Started", 0, "Low", "3", "Not implemented"],

    # 19. Permission Management
    ["19. Permission Management", "Permission Structure", "Complete", 100, "High", "1", "Permissions table"],
    ["19. Permission Management", "Role-Level Permissions", "Complete", 100, "High", "1", "RolePermissions table"],
    ["19. Permission Management", "User-Level Permissions", "Complete", 100, "High", "1", "UserPermissions table"],
    ["19. Permission Management", "Field-Level Permissions", "Not Started", 0, "High", "2", "Not implemented"],
    ["19. Permission Management", "Record-Level Permissions", "Not Started", 0, "High", "2", "Not implemented"],
    ["19. Permission Management", "Time-Based Permissions", "Not Started", 0, "Low", "3", "Not implemented"],
    ["19. Permission Management", "Permission Delegation", "Not Started", 0, "Low", "3", "Not implemented"],

    # 20. Authentication & Security
    ["20. Authentication & Security", "Email/Password Authentication", "Complete", 100, "High", "1", "JWT-based"],
    ["20. Authentication & Security", "JWT Token Authentication", "Complete", 100, "High", "1", "Implemented"],
    ["20. Authentication & Security", "Refresh Token Mechanism", "Complete", 100, "High", "1", "Implemented"],
    ["20. Authentication & Security", "Email Verification", "Complete", 100, "High", "1", "With tokens"],
    ["20. Authentication & Security", "Password Reset", "Complete", 100, "High", "1", "With expiration"],
    ["20. Authentication & Security", "Account Locking", "Complete", 100, "High", "1", "Brute-force protection"],
    ["20. Authentication & Security", "Two-Factor Authentication (2FA)", "Not Started", 0, "High", "2", "Not implemented"],
    ["20. Authentication & Security", "OAuth Integration", "Not Started", 0, "Medium", "2", "Not implemented"],
    ["20. Authentication & Security", "SSO/SAML Integration", "Not Started", 0, "Medium", "2", "Not implemented"],
    ["20. Authentication & Security", "LDAP/AD Integration", "Not Started", 0, "Medium", "2", "Not implemented"],
    ["20. Authentication & Security", "Password Strength Enforcement", "Not Started", 0, "Medium", "2", "Not implemented"],
    ["20. Authentication & Security", "Password History", "Not Started", 0, "Low", "3", "Not implemented"],

    # 21. Audit & Compliance
    ["21. Audit & Compliance", "Audit Trail", "Complete", 100, "High", "1", "AuditLogs table"],
    ["21. Audit & Compliance", "Security Logs", "Complete", 100, "High", "1", "SecurityLogs table"],
    ["21. Audit & Compliance", "Login/Logout Tracking", "Complete", 100, "High", "1", "Implemented"],
    ["21. Audit & Compliance", "Entity-Level Audit Fields", "Complete", 100, "High", "1", "CreatedBy, UpdatedBy"],
    ["21. Audit & Compliance", "Detailed Change Tracking", "Not Started", 0, "Medium", "2", "Before/after values"],
    ["21. Audit & Compliance", "Audit Log Search & Filter", "Not Started", 0, "Medium", "2", "Not implemented"],
    ["21. Audit & Compliance", "GDPR Compliance Tools", "Not Started", 0, "Medium", "2", "Not implemented"],
    ["21. Audit & Compliance", "Data Retention Policies", "Not Started", 0, "Low", "3", "Not implemented"],

    # 22. Notifications & Communication
    ["22. Notifications", "In-App Notifications", "Complete", 100, "High", "1", "Notifications table"],
    ["22. Notifications", "Notification Preferences", "Complete", 100, "Medium", "1", "Preferences exist"],
    ["22. Notifications", "Email Notifications", "Not Started", 0, "Medium", "2", "Not implemented"],
    ["22. Notifications", "SMS Notifications", "Not Started", 0, "Low", "3", "Not implemented"],
    ["22. Notifications", "Push Notifications", "Not Started", 0, "Low", "3", "Not implemented"],
    ["22. Notifications", "Real-Time Notifications", "Not Started", 0, "Medium", "2", "WebSocket not implemented"],
]

# Write data
for row_idx, row_data in enumerate(data, 5):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        
        # Color coding based on status
        if col_idx == 3:  # Status column
            if value == "Complete":
                cell.fill = complete_fill
            elif value == "Partial":
                cell.fill = partial_fill
            elif value == "Not Started":
                cell.fill = not_started_fill

# Calculate and add summary section
summary_row = len(data) + 7
ws.merge_cells(f'A{summary_row}:G{summary_row}')
ws[f'A{summary_row}'] = "SUMMARY STATISTICS"
ws[f'A{summary_row}'].font = Font(size=14, bold=True)
ws[f'A{summary_row}'].alignment = Alignment(horizontal='center', vertical='center')

# Calculate statistics
total_features = len(data)
complete_count = sum(1 for row in data if row[2] == "Complete")
partial_count = sum(1 for row in data if row[2] == "Partial")
not_started_count = sum(1 for row in data if row[2] == "Not Started")

complete_percentage = round((complete_count / total_features) * 100, 1)
partial_percentage = round((partial_count / total_features) * 100, 1)
not_started_percentage = round((not_started_count / total_features) * 100, 1)
overall_completion = round((complete_count + (partial_count * 0.5)) / total_features * 100, 1)

summary_data = [
    ["Total Features", total_features, "", "", "", "", ""],
    ["Complete", complete_count, f"{complete_percentage}%", "", "", "", ""],
    ["Partial", partial_count, f"{partial_percentage}%", "", "", "", ""],
    ["Not Started", not_started_count, f"{not_started_percentage}%", "", "", "", ""],
    ["Overall Completion (Weighted)", "", f"{overall_completion}%", "", "", "", ""],
    ["Remaining Work", "", f"{100 - overall_completion}%", "", "", "", ""],
]

for row_idx, row_data in enumerate(summary_data, summary_row + 1):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.border = thin_border
        if col_idx in [1, 3]:
            cell.font = Font(bold=True)

# Module-wise summary
module_summary_row = summary_row + 9
ws.merge_cells(f'A{module_summary_row}:G{module_summary_row}')
ws[f'A{module_summary_row}'] = "MODULE-WISE COMPLETION"
ws[f'A{module_summary_row}'].font = Font(size=14, bold=True)
ws[f'A{module_summary_row}'].alignment = Alignment(horizontal='center', vertical='center')

# Calculate module-wise completion
modules = {}
for row in data:
    module = row[0]
    if module not in modules:
        modules[module] = {"total": 0, "complete": 0, "partial": 0}
    modules[module]["total"] += 1
    if row[2] == "Complete":
        modules[module]["complete"] += 1
    elif row[2] == "Partial":
        modules[module]["partial"] += 1

# Module summary headers
module_headers = ["Module", "Total Features", "Complete", "Partial", "Not Started", "Completion %", "Status"]
for col, header in enumerate(module_headers, 1):
    cell = ws.cell(row=module_summary_row + 1, column=col, value=header)
    cell.fill = header_fill
    cell.font = header_font
    cell.border = thin_border
    cell.alignment = Alignment(horizontal='center', vertical='center')

# Write module summary
row_num = module_summary_row + 2
for module, stats in modules.items():
    not_started = stats["total"] - stats["complete"] - stats["partial"]
    completion = round((stats["complete"] + (stats["partial"] * 0.5)) / stats["total"] * 100, 1)
    
    status = "Complete" if completion == 100 else "Partial" if completion > 0 else "Not Started"
    
    module_data = [
        module,
        stats["total"],
        stats["complete"],
        stats["partial"],
        not_started,
        f"{completion}%",
        status
    ]
    
    for col_idx, value in enumerate(module_data, 1):
        cell = ws.cell(row=row_num, column=col_idx, value=value)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        
        if col_idx == 7:  # Status column
            if value == "Complete":
                cell.fill = complete_fill
            elif value == "Partial":
                cell.fill = partial_fill
            elif value == "Not Started":
                cell.fill = not_started_fill
    
    row_num += 1

# Adjust column widths
column_widths = [35, 45, 12, 15, 12, 10, 40]
for i, width in enumerate(column_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = width

# Freeze panes
ws.freeze_panes = 'A5'

# Save workbook
output_file = "ERP_CRM_Project_Completion_Status.xlsx"
wb.save(output_file)
print(f"✅ Excel file created: {output_file}")
print(f"\n📊 Project Statistics:")
print(f"   Total Features: {total_features}")
print(f"   Complete: {complete_count} ({complete_percentage}%)")
print(f"   Partial: {partial_count} ({partial_percentage}%)")
print(f"   Not Started: {not_started_count} ({not_started_percentage}%)")
print(f"   Overall Completion: {overall_completion}%")
print(f"   Remaining Work: {100 - overall_completion}%")